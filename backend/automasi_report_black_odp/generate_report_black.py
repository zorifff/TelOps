import pandas as pd
import openpyxl
from copy import copy
import re
import os

woks_in_template = [
    "KEBUMEN", "MAGELANG TEMANGGUNG", "BATANG", "PEMALANG PURBALINGGA", 
    "TEGAL BREBES", "CILACAP BANYUMAS", "WONOSOBO BANJARNEGARA", "DEMAK", 
    "JEPARA KUDUS - PATI", "SEMARANG 1", "SEMARANG 2", "BOYOLALI", 
    "SRAGEN", "SURAKARTA", "YOGYA 1", "YOGYA 2"
]

durations = ['<1 Month', '<2 Month', '<3 Month', '4-6 Month', '>6 Month']

branches_map_template = {
    "MAGELANG": range(0, 2),
    "PEKALONGAN": range(2, 5),
    "PURWOKERTO": range(5, 7),
    "SEMARANG": range(7, 11),
    "SURAKARTA": range(11, 14),
    "YOGYAKARTA": range(14, 16)
}

def duplicate_cf_for_block(ws, offset):
    """
    Duplicate the 16 template Conditional Formatting rules for a new table block at offset.
    Update row numbers in formulas so formulas reference cells in their own table.
    """
    template_cfs = list(ws.conditional_formatting)[:16]
    for cf in template_cfs:
        sqref_str = str(cf.sqref)
        ranges = sqref_str.split(' ')
        new_ranges = []
        for r_str in ranges:
            if ':' in r_str:
                start_c, end_c = r_str.split(':')
                c_start_let = ''.join(filter(str.isalpha, start_c))
                r_start_num = int(''.join(filter(str.isdigit, start_c))) + offset
                c_end_let = ''.join(filter(str.isalpha, end_c))
                r_end_num = int(''.join(filter(str.isdigit, end_c))) + offset
                new_ranges.append(f'{c_start_let}{r_start_num}:{c_end_let}{r_end_num}')
            else:
                c_let = ''.join(filter(str.isalpha, r_str))
                r_num = int(''.join(filter(str.isdigit, r_str))) + offset
                new_ranges.append(f'{c_let}{r_num}')
        
        new_sqref = ' '.join(new_ranges)
        
        for rule in cf.rules:
            new_rule = copy(rule)
            if new_rule.formula:
                new_formulas = []
                for f in new_rule.formula:
                    def replace_row(match):
                        col = match.group(1)
                        r_num = int(match.group(2)) + offset
                        return f'{col}{r_num}'
                    
                    f_updated = re.sub(r'([A-Z]+)(\d+)', replace_row, str(f))
                    new_formulas.append(f_updated)
                new_rule.formula = new_formulas
            
            ws.conditional_formatting.add(new_sqref, new_rule)

def copy_template_table(ws, src_start=2, src_end=30, dest_start=33, title_text=''):
    """
    Duplicate Table 1 layout, styles, row heights, merged cells & conditional formatting to dest_start.
    """
    offset = dest_start - src_start

    # Copy row heights
    for r in range(src_start, src_end + 1):
        if r in ws.row_dimensions and ws.row_dimensions[r].height is not None:
            ws.row_dimensions[r + offset].height = ws.row_dimensions[r].height

    # Copy cell styles and content
    for r in range(src_start, src_end + 1):
        for c in range(1, 27):
            src_cell = ws.cell(row=r, column=c)
            dest_cell = ws.cell(row=r + offset, column=c)
            
            dest_cell.value = src_cell.value
            if src_cell.has_style:
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.protection = copy(src_cell.protection)
                dest_cell.alignment = copy(src_cell.alignment)

    # Copy merged cell ranges within the block
    merged_to_add = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row >= src_start and rng.max_row <= src_end:
            new_min_row = rng.min_row + offset
            new_max_row = rng.max_row + offset
            new_range = f'{openpyxl.utils.get_column_letter(rng.min_col)}{new_min_row}:{openpyxl.utils.get_column_letter(rng.max_col)}{new_max_row}'
            merged_to_add.append(new_range)
            
    for m in merged_to_add:
        ws.merge_cells(m)

    # Duplicate Conditional Formatting for this block
    duplicate_cf_for_block(ws, offset)

    if title_text:
        ws.cell(row=dest_start, column=2).value = title_text

def extract_w1_table_data(excel_w1, w1_path, td_filter):
    """
    Extract W-1 Black ODP previous week data strictly for the specified td_filter ('GREENFIELD', 'BROWNFIELD', or 'ALL').
    If the specified table does not exist in W-1's Report - ODP Black sheet, returns zeros/nulls.
    """
    w1_black = {wok: 0 for wok in woks_in_template}
    w1_total = {wok: 0 for wok in woks_in_template}

    if "Report - ODP Black" in excel_w1.sheet_names:
        df_w1 = pd.read_excel(w1_path, sheet_name="Report - ODP Black", header=None)

        table_start_r = -1
        target_keyword = td_filter.upper()
        if target_keyword in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
            target_keyword = 'ALL'

        for r in range(len(df_w1)):
            row_str = ' '.join(str(df_w1.iloc[r, c]).strip().upper() for c in range(min(5, df_w1.shape[1])) if pd.notna(df_w1.iloc[r, c]))
            if 'REPORT ODP BLACK' in row_str or 'BLACK ODP' in row_str:
                if target_keyword in row_str:
                    table_start_r = r
                    break

        # Fallback for Greenfield if title line doesn't explicitly contain 'GREENFIELD'
        if table_start_r == -1 and target_keyword == 'GREENFIELD':
            table_start_r = 1  # Row 2 (0-indexed 1)

        if table_start_r != -1:
            wok_start_r = table_start_r + 5  # Row 7 (0-indexed 6) when title is at row 2 (0-indexed 1)
            for r in range(wok_start_r, min(wok_start_r + 20, len(df_w1))):
                row = df_w1.iloc[r]
                wok_name = str(row[1]).strip() if len(row) > 1 and pd.notna(row[1]) else ""
                if wok_name in woks_in_template:
                    val_b = row[18] if len(row) > 18 else None
                    if val_b is None or pd.isna(val_b) or str(val_b).strip() in ('#N/A', 'None', 'nan', ''):
                        val_b = row[20] if len(row) > 20 else None
                    try:
                        if pd.notna(val_b):
                            w1_black[wok_name] = int(float(val_b))
                    except:
                        pass

                    val_t = row[17] if len(row) > 17 else None
                    try:
                        if pd.notna(val_t):
                            w1_total[wok_name] = int(float(val_t))
                    except:
                        pass
            return w1_black, w1_total
        else:
            # Table for td_filter NOT found in Report - ODP Black! Return zeros so Previous Week is empty/null!
            print(f"Informasi: Tabel {td_filter} tidak ditemukan pada sheet Report - ODP Black file W-1. Previous week diset kosong.")
            return w1_black, w1_total

    elif "ODP Golive 2026" in excel_w1.sheet_names:
        df_raw = pd.read_excel(w1_path, sheet_name="ODP Golive 2026")
        used_col = 'Used_new_v3' if 'Used_new_v3' in df_raw.columns else 'Used_new_v2'
        
        df_raw['TD_Clean'] = df_raw['Type Design'].astype(str).str.strip().str.upper()
        df_raw['OCC2_Clean'] = df_raw['OCC 2'].astype(str).str.strip().str.upper()
        df_raw['WOK_Clean'] = df_raw['WOK'].astype(str).str.strip()
        df_raw[used_col] = pd.to_numeric(df_raw[used_col], errors='coerce').fillna(0)
        df_raw.loc[df_raw[used_col] < 0, used_col] = 0

        if td_filter.upper() not in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
            mask_tot = df_raw['TD_Clean'] == td_filter.upper()
        else:
            mask_tot = pd.Series(True, index=df_raw.index)

        mask_blk = mask_tot & (df_raw[used_col] == 0) & (df_raw['OCC2_Clean'] == 'BLACK')

        agg_tot = df_raw[mask_tot].groupby('WOK_Clean').size()
        agg_blk = df_raw[mask_blk].groupby('WOK_Clean').size()

        for wok in woks_in_template:
            w1_black[wok] = int(agg_blk.get(wok, 0))
            w1_total[wok] = int(agg_tot.get(wok, 0))

    return w1_black, w1_total

def process_table(ws_out, df_w0, excel_w1, w1_path, td_filter, start_wok_row=7):
    used_col = 'Used_new_v3' if 'Used_new_v3' in df_w0.columns else 'Used_new_v2'
    
    df_w0_sub = df_w0.copy()
    df_w0_sub['TD_Clean'] = df_w0_sub['Type Design'].astype(str).str.strip().str.upper()
    df_w0_sub['OCC2_Clean'] = df_w0_sub['OCC 2'].astype(str).str.strip().str.upper()
    df_w0_sub['Dur_Clean'] = df_w0_sub['Cat Durasi Go Live'].astype(str).str.strip()
    df_w0_sub[used_col] = pd.to_numeric(df_w0_sub[used_col], errors='coerce').fillna(0)
    df_w0_sub.loc[df_w0_sub[used_col] < 0, used_col] = 0

    td_clean = td_filter.strip().upper()
    if td_clean not in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
        mask_tot = df_w0_sub['TD_Clean'] == td_clean
    else:
        mask_tot = pd.Series(True, index=df_w0_sub.index)

    mask_blk = mask_tot & (df_w0_sub[used_col] == 0) & (df_w0_sub['OCC2_Clean'] == 'BLACK')

    df_tot = df_w0_sub[mask_tot]
    df_blk = df_w0_sub[mask_blk]

    agg_tot = df_tot.groupby(['WOK', 'Dur_Clean']).size().reset_index(name='Total_ODP')
    agg_blk = df_blk.groupby(['WOK', 'Dur_Clean']).size().reset_index(name='Black_ODP')

    wok_data = {wok: {d: {'Total': 0, 'Black': 0} for d in durations} for wok in woks_in_template}

    for _, row in agg_tot.iterrows():
        wok = str(row['WOK']).strip()
        dur = str(row['Dur_Clean']).strip()
        if wok in wok_data and dur in wok_data[wok]:
            wok_data[wok][dur]['Total'] = row['Total_ODP']

    for _, row in agg_blk.iterrows():
        wok = str(row['WOK']).strip()
        dur = str(row['Dur_Clean']).strip()
        if wok in wok_data and dur in wok_data[wok]:
            wok_data[wok][dur]['Black'] = row['Black_ODP']

    w1_black, w1_total = extract_w1_table_data(excel_w1, w1_path, td_clean)

    row_data = {}
    cols = ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P']

    # WOK Rows
    for idx, wok in enumerate(woks_in_template):
        r = start_wok_row + idx
        d = wok_data[wok]

        vals = [d['<1 Month']['Total'], d['<1 Month']['Black'],
                d['<2 Month']['Total'], d['<2 Month']['Black'],
                d['<3 Month']['Total'], d['<3 Month']['Black'],
                d['4-6 Month']['Total'], d['4-6 Month']['Black'],
                d['>6 Month']['Total'], d['>6 Month']['Black']]

        for c, v in zip(cols, vals):
            ws_out[f'{c}{r}'] = v

        ws_out[f'E{r}'] = d['<1 Month']['Black'] / d['<1 Month']['Total'] if d['<1 Month']['Total'] > 0 else '-'
        ws_out[f'H{r}'] = d['<2 Month']['Black'] / d['<2 Month']['Total'] if d['<2 Month']['Total'] > 0 else '-'
        ws_out[f'K{r}'] = d['<3 Month']['Black'] / d['<3 Month']['Total'] if d['<3 Month']['Total'] > 0 else '-'
        ws_out[f'N{r}'] = d['4-6 Month']['Black'] / d['4-6 Month']['Total'] if d['4-6 Month']['Total'] > 0 else '-'
        ws_out[f'Q{r}'] = d['>6 Month']['Black'] / d['>6 Month']['Total'] if d['>6 Month']['Total'] > 0 else '-'

        tot_all = sum(v for i, v in enumerate(vals) if i % 2 == 0)
        blk_all = sum(v for i, v in enumerate(vals) if i % 2 == 1)

        ws_out[f'R{r}'] = tot_all
        ws_out[f'S{r}'] = blk_all
        ws_out[f'T{r}'] = blk_all / tot_all if tot_all > 0 else '-'

        w1_blk = w1_black[wok]
        w1_tot = w1_total[wok]

        if w1_tot > 0:
            ws_out[f'U{r}'] = w1_blk
            w1_pct = w1_blk / w1_tot
            ws_out[f'V{r}'] = w1_pct
        else:
            ws_out[f'U{r}'] = "-"
            ws_out[f'V{r}'] = "-"
            w1_pct = 0.0

        w_diff = blk_all - w1_blk
        ws_out[f'W{r}'] = w_diff

        t_val = (blk_all / tot_all) if tot_all > 0 else 0
        x_diff = t_val - w1_pct
        ws_out[f'X{r}'] = x_diff

        ws_out[f'Y{r}'] = 'Berkurang' if w_diff < 0 else 'Bertambah' if w_diff > 0 else '-'
        ws_out[f'Z{r}'] = 'Berkurang' if x_diff < 0 else 'Bertambah' if x_diff > 0 else '-'

        row_data[r] = {
            'C': vals[0], 'D': vals[1], 'F': vals[2], 'G': vals[3],
            'I': vals[4], 'J': vals[5], 'L': vals[6], 'M': vals[7],
            'O': vals[8], 'P': vals[9], 'R': tot_all, 'S': blk_all,
            'U': w1_blk, 'W1_Total': w1_tot
        }

    # Branch Rows
    start_branch_row = start_wok_row + 17
    branch_rows_written = []
    for b_idx, (branch, wok_indices) in enumerate(branches_map_template.items()):
        dest = start_branch_row + b_idx
        branch_rows_written.append(dest)
        
        sum_cols = {c: 0 for c in ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S', 'U', 'W1_Total']}
        for w_i in wok_indices:
            r = start_wok_row + w_i
            for c in sum_cols:
                sum_cols[c] += row_data[r][c]

        for c in ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S']:
            ws_out[f'{c}{dest}'] = sum_cols[c]

        if sum_cols['W1_Total'] > 0:
            ws_out[f'U{dest}'] = sum_cols['U']
            w1_pct = sum_cols['U'] / sum_cols['W1_Total']
            ws_out[f'V{dest}'] = w1_pct
        else:
            ws_out[f'U{dest}'] = "-"
            ws_out[f'V{dest}'] = "-"
            w1_pct = 0.0

        ws_out[f'E{dest}'] = sum_cols['D'] / sum_cols['C'] if sum_cols['C'] > 0 else '-'
        ws_out[f'H{dest}'] = sum_cols['G'] / sum_cols['F'] if sum_cols['F'] > 0 else '-'
        ws_out[f'K{dest}'] = sum_cols['J'] / sum_cols['I'] if sum_cols['I'] > 0 else '-'
        ws_out[f'N{dest}'] = sum_cols['M'] / sum_cols['L'] if sum_cols['L'] > 0 else '-'
        ws_out[f'Q{dest}'] = sum_cols['P'] / sum_cols['O'] if sum_cols['O'] > 0 else '-'
        ws_out[f'T{dest}'] = sum_cols['S'] / sum_cols['R'] if sum_cols['R'] > 0 else '-'

        w_diff = sum_cols['S'] - sum_cols['U']
        ws_out[f'W{dest}'] = w_diff

        t_val = sum_cols['S'] / sum_cols['R'] if sum_cols['R'] > 0 else 0
        x_diff = t_val - w1_pct
        ws_out[f'X{dest}'] = x_diff

        ws_out[f'Y{dest}'] = 'Berkurang' if w_diff < 0 else 'Bertambah' if w_diff > 0 else '-'
        ws_out[f'Z{dest}'] = 'Berkurang' if x_diff < 0 else 'Bertambah' if x_diff > 0 else '-'

        row_data[dest] = sum_cols

    # Total Row (Jateng DIY)
    tot_dest = start_branch_row + len(branches_map_template)
    sum_cols = {c: 0 for c in ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S', 'U', 'W1_Total']}
    for r in branch_rows_written:
        for c in sum_cols:
            sum_cols[c] += row_data[r][c]

    for c in ['C', 'D', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S']:
        ws_out[f'{c}{tot_dest}'] = sum_cols[c]

    if sum_cols['W1_Total'] > 0:
        ws_out[f'U{tot_dest}'] = sum_cols['U']
        w1_pct = sum_cols['U'] / sum_cols['W1_Total']
        ws_out[f'V{tot_dest}'] = w1_pct
    else:
        ws_out[f'U{tot_dest}'] = "-"
        ws_out[f'V{tot_dest}'] = "-"
        w1_pct = 0.0

    ws_out[f'E{tot_dest}'] = sum_cols['D'] / sum_cols['C'] if sum_cols['C'] > 0 else '-'
    ws_out[f'H{tot_dest}'] = sum_cols['G'] / sum_cols['F'] if sum_cols['F'] > 0 else '-'
    ws_out[f'K{tot_dest}'] = sum_cols['J'] / sum_cols['I'] if sum_cols['I'] > 0 else '-'
    ws_out[f'N{tot_dest}'] = sum_cols['M'] / sum_cols['L'] if sum_cols['L'] > 0 else '-'
    ws_out[f'Q{tot_dest}'] = sum_cols['P'] / sum_cols['O'] if sum_cols['O'] > 0 else '-'
    ws_out[f'T{tot_dest}'] = sum_cols['S'] / sum_cols['R'] if sum_cols['R'] > 0 else '-'

    w_diff = sum_cols['S'] - sum_cols['U']
    ws_out[f'W{tot_dest}'] = w_diff

    t_val = sum_cols['S'] / sum_cols['R'] if sum_cols['R'] > 0 else 0
    x_diff = t_val - w1_pct
    ws_out[f'X{tot_dest}'] = x_diff

    ws_out[f'Y{tot_dest}'] = 'Berkurang' if w_diff < 0 else 'Bertambah' if w_diff > 0 else '-'
    ws_out[f'Z{tot_dest}'] = 'Berkurang' if x_diff < 0 else 'Bertambah' if x_diff > 0 else '-'

def generate_report_black(w0_path, w1_path, out_path, template_path, td_filter='GREENFIELD'):
    if not os.path.exists(w0_path):
        raise FileNotFoundError("Berkas W-0 (Raw Data) tidak ditemukan atau telah dihapus.")
    if not os.path.exists(w1_path):
        raise FileNotFoundError("Berkas W-1 (Minggu Lalu) tidak ditemukan atau telah dihapus.")
    if not os.path.exists(template_path):
        raise FileNotFoundError("Berkas Template_Report_Black.xlsx tidak ditemukan di server.")

    # --- VALIDASI BERKAS W-0 ---
    try:
        with pd.ExcelFile(w0_path) as excel_w0:
            if "ODP Golive 2026" not in excel_w0.sheet_names:
                raise ValueError("Berkas W-0 (Raw Data) tidak memiliki sheet 'ODP Golive 2026'. Harap periksa kembali berkas Anda.")
        df_w0 = pd.read_excel(w0_path, sheet_name="ODP Golive 2026")
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise ValueError(f"Gagal membaca berkas W-0 (Raw Data): {e}")

    required_cols_w0 = ['WOK', 'Type Design', 'OCC 2', 'Cat Durasi Go Live']
    missing_w0 = [c for c in required_cols_w0 if c not in df_w0.columns]
    if missing_w0:
        raise ValueError(f"Berkas W-0 (Raw Data) tidak memiliki kolom wajib berikut: {', '.join(missing_w0)}. Harap periksa panduan file.")

    # --- READ W-1 EXCEL FILE ---
    try:
        excel_w1 = pd.ExcelFile(w1_path)
    except Exception as e:
        raise ValueError(f"Gagal membaca berkas W-1 (Minggu Lalu): {e}")

    wb_out = openpyxl.load_workbook(template_path)
    try:
        if "Report - ODP Black" not in wb_out.sheetnames:
            raise ValueError("Sheet 'Report - ODP Black' tidak ditemukan pada template.")
        ws_out = wb_out["Report - ODP Black"]

        td_clean = td_filter.strip().upper()

        # Check if user requested COMBINED mode (All 3 tables on 1 sheet: Greenfield, Brownfield, All Type Design)
        if td_clean in ('COMBINED', 'COMBINED_ALL', 'ALL_TABLES', '3_TABLES'):
            # Table 1: Greenfield (Title at Row 2, WOKs start at Row 7)
            ws_out['B2'] = "Report ODP Black Golive 2026 - Greenfield"
            process_table(ws_out, df_w0, excel_w1, w1_path, 'GREENFIELD', start_wok_row=7)

            # Table 2: Brownfield (Title at Row 33, WOKs start at Row 38)
            copy_template_table(ws_out, src_start=2, src_end=30, dest_start=33, title_text="Report ODP Black Golive 2026 - Brownfield")
            process_table(ws_out, df_w0, excel_w1, w1_path, 'BROWNFIELD', start_wok_row=38)

            # Table 3: All Type Design (Title at Row 64, WOKs start at Row 69)
            copy_template_table(ws_out, src_start=2, src_end=30, dest_start=64, title_text="Report ODP Black Golive 2026 - All Type Design")
            process_table(ws_out, df_w0, excel_w1, w1_path, 'ALL', start_wok_row=69)

        else:
            # Single-table mode (Greenfield, Brownfield, or All Type Design)
            if td_clean == 'BROWNFIELD':
                ws_out['B2'] = "Report ODP Black Golive 2026 - Brownfield"
            elif td_clean in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
                ws_out['B2'] = "Report ODP Black Golive 2026 - All Type Design"
            else:
                ws_out['B2'] = "Report ODP Black Golive 2026 - Greenfield"

            process_table(ws_out, df_w0, excel_w1, w1_path, td_clean, start_wok_row=7)

        wb_out.save(out_path)
    finally:
        try:
            excel_w1.close()
        except:
            pass
        try:
            wb_out.close()
        except Exception:
            pass
