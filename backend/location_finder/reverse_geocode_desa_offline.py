#!/usr/bin/env python3
"""Cari alamat administratif dari titik Excel dengan polygon desa/kelurahan.

Tidak menggunakan API maupun jaringan. Data polygon yang dipakai adalah
shapefile di dalam ZIP ``DESA-KECAMATAN JATENG DIY.zip`` (WGS84 / EPSG:4326).

Contoh:
  python reverse_geocode_desa_offline.py "raw order 18 jul.xlsx" \
    "DESA-KECAMATAN JATENG DIY.zip" hasil_alamat.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import shapefile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from shapely.geometry import Point, shape
from shapely.strtree import STRtree


REQUIRED_BOUNDARY_FIELDS = ("PROVINSI", "KAB_KOTA", "KECAMATAN", "DESA_KELUR")


def normalise_header(value: object) -> str:
    return str(value or "").strip().lower().replace("_", "").replace(" ", "")


def find_lat_lng_cols(ws) -> tuple[int, int, list[str]]:
    headers = [str(cell.value or "") if cell.value is not None else "" for cell in ws[1]]
    norm_headers = [normalise_header(h) for h in headers]
    
    lat_col = -1
    lng_col = -1
    
    for i, h in enumerate(norm_headers, start=1):
        if h in ('latitude', 'lat'):
            lat_col = i
        elif h in ('longitude', 'long', 'lng'):
            lng_col = i
            
    if lat_col == -1 or lng_col == -1:
        raise ValueError("Kolom Latitude/Lat dan Longitude/Lng/Long wajib ada.")
    return lat_col, lng_col, headers


def title_case(value: object) -> str:
    """Rapikan nilai huruf kapital dari shapefile, tanpa mengubah singkatan."""
    text = str(value or "").strip()
    return text.title() if text.isupper() else text


def administrative_address(record: dict[str, object]) -> tuple[str, str, str, str]:
    """Buat komponen alamat dari atribut polygon yang memuat titik tersebut."""
    desa = title_case(record["DESA_KELUR"])
    kecamatan = title_case(record["KECAMATAN"])
    kab_kota = title_case(record["KAB_KOTA"])
    
    upper_kab = kab_kota.upper()
    stripped_kab = upper_kab
    if upper_kab.startswith("KABUPATEN "):
        stripped_kab = upper_kab[10:].strip()
    elif upper_kab.startswith("KOTA "):
        stripped_kab = upper_kab[5:].strip()
        
    if stripped_kab not in ["MAGELANG", "SEMARANG", "TEGAL", "PEKALONGAN"]:
        if upper_kab.startswith("KABUPATEN "):
            kab_kota = kab_kota[10:].strip()
        elif upper_kab.startswith("KOTA "):
            kab_kota = kab_kota[5:].strip()

    provinsi = title_case(record["PROVINSI"])
    return desa, kecamatan, kab_kota, provinsi


def load_boundaries(zip_path: Path):
    """Baca shapefile ZIP tanpa mengekstrak file ke disk."""
    with ZipFile(zip_path) as archive:
        shp_files = [name for name in archive.namelist() if name.lower().endswith(".shp")]
        if len(shp_files) != 1:
            raise ValueError("ZIP harus memiliki tepat satu file .shp")
        stem = shp_files[0][:-4]
        expected = [stem + suffix for suffix in (".shp", ".shx", ".dbf")]
        absent = [name for name in expected if name not in archive.namelist()]
        if absent:
            raise ValueError("Komponen shapefile tidak lengkap: " + ", ".join(absent))
        reader = shapefile.Reader(
            shp=BytesIO(archive.read(stem + ".shp")),
            shx=BytesIO(archive.read(stem + ".shx")),
            dbf=BytesIO(archive.read(stem + ".dbf")),
            encoding="utf-8",
        )
        names = [field[0] for field in reader.fields[1:]]
        missing = set(REQUIRED_BOUNDARY_FIELDS) - set(names)
        if missing:
            raise ValueError("Kolom administrasi tidak ditemukan: " + ", ".join(sorted(missing)))
        geometries, records = [], []
        for feature in reader.iterShapeRecords():
            geometry = shape(feature.shape.__geo_interface__)
            if geometry.is_empty:
                continue
            record = dict(zip(names, feature.record))
            geometries.append(geometry)
            records.append(record)
    if not geometries:
        raise ValueError("Tidak ada polygon yang dapat dibaca dari shapefile")
    return geometries, records, STRtree(geometries)


def candidate_indices(tree: STRtree, point: Point) -> list[int]:
    """Kompatibel dengan STRtree Shapely 2.x (mengembalikan indeks)."""
    values = tree.query(point)
    return [int(value) for value in values]


def lookup_address(point: Point, geometries, records, tree) -> tuple[tuple[str, str, str, str], str]:
    """Gunakan covers agar titik tepat pada garis batas tetap terdeteksi.

    Bila lebih dari satu polygon mencakup titik, hasil sengaja tidak dipilih
    agar alamat tidak keliru. Pengguna dapat meninjau baris tersebut di Log.
    """
    matched = [index for index in candidate_indices(tree, point) if geometries[index].covers(point)]
    if len(matched) == 1:
        return administrative_address(records[matched[0]]), "OK"
    if not matched:
        return ("", "", "", ""), "TIDAK_DITEMUKAN_DI_BATAS_DATA"
    return ("", "", "", ""), "AMBIGU_DI_GARIS_ATAU_OVERLAP_BATAS"


def style_workbook(result, log, num_original_cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for ws in (result, log):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")


def process_offline(input_xlsx: Path, boundary_zip: Path, output_xlsx: Path) -> dict:
    if not input_xlsx.is_file() or not boundary_zip.is_file():
        raise FileNotFoundError("File input Excel atau ZIP batas tidak ditemukan.")
    if input_xlsx.resolve() == output_xlsx.resolve():
        raise ValueError("Output harus memakai nama yang berbeda dari file sumber.")

    try:
        geometries, records, tree = load_boundaries(boundary_zip)
        source = load_workbook(input_xlsx, read_only=True, data_only=True)
        source_ws = source[source.sheetnames[0]]
        lat_col, lng_col, headers = find_lat_lng_cols(source_ws)
    except Exception as exc:
        raise RuntimeError(f"Gagal membaca data: {exc}")

    out = Workbook()
    result = out.active
    result.title = "Hasil"
    log = out.create_sheet("Log")
    result.append(headers + ["desa", "kecamatan", "kabupaten", "provinsi"])
    log.append(["baris_sumber", "latitude", "longitude", "status", "catatan"])
    counts: dict[str, int] = {}

    for row_number, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
        longitude_raw = row[lng_col - 1]
        latitude_raw = row[lat_col - 1]
        try:
            longitude, latitude = float(longitude_raw), float(latitude_raw)
            if not (-180 <= longitude <= 180 and -90 <= latitude <= 90):
                raise ValueError("koordinat di luar rentang WGS84")
            address_tuple, status = lookup_address(Point(longitude, latitude), geometries, records, tree)
            note = "" if status == "OK" else "Tinjau koordinat atau kelengkapan polygon batas desa."
        except (TypeError, ValueError) as exc:
            address_tuple, status, note = ("", "", "", ""), "KOORDINAT_TIDAK_VALID", str(exc)
        result.append(list(row) + list(address_tuple))
        log.append([row_number, latitude_raw, longitude_raw, status, note])
        counts[status] = counts.get(status, 0) + 1

    style_workbook(result, log, len(headers))
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out.save(output_xlsx)
    return counts
