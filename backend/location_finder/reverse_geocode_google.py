#!/usr/bin/env python3
"""Reverse-geocode koordinat Excel memakai Google Maps Geocoding API.

Contoh:
  set GOOGLE_MAPS_API_KEY=AIza...
  python reverse_geocode_google.py "raw order 18 jul.xlsx" hasil_alamat.xlsx
"""
from __future__ import annotations
import argparse, os, random, sys, time
from collections import defaultdict
from pathlib import Path
from typing import Any
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

API_URL = "https://maps.googleapis.com/maps/api/geocode/json"
VILLAGE_TYPES = ("administrative_area_level_4", "sublocality_level_1", "sublocality", "locality")

def normalise_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", "").replace(" ", "")

def find_lat_lng_cols(ws) -> tuple[int, int, list[str]]:
    headers = [str(cell.value or "") if cell.value is not None else "" for cell in ws[1]]
    norm_headers = [normalise_header(h) for h in headers]
    
    lat_col = -1
    lng_col = -1
    
    for i, h in enumerate(norm_headers, start=1):
        if h in ('latitude', 'lat'):
            lat_col = i
        elif h in ('longitude', 'long', 'lng'):
            lng_col = i
            
    if lat_col == -1 or lng_col == -1:
        raise ValueError("Kolom Latitude/Lat dan Longitude/Lng/Long wajib ada.")
    return lat_col, lng_col, headers

def component_index(results: list[dict[str, Any]]) -> dict[str, str]:
    """Mengambil komponen Google, tanpa mengarang wilayah yang tidak ada."""
    found: dict[str, str] = {}
    for result in results:
        for comp in result.get("address_components", []):
            for kind in comp.get("types", []):
                found.setdefault(kind, comp.get("long_name", "").strip())
    return found

def structured_address(results: list[dict[str, Any]]) -> tuple[tuple[str, str, str, str], bool, str]:
    components = component_index(results)
    village = next((components[t] for t in VILLAGE_TYPES if components.get(t)), "")
    district = components.get("administrative_area_level_3", "")
    regency = components.get("administrative_area_level_2", "")
    province = components.get("administrative_area_level_1", "")
    
    upper_regency = regency.upper()
    stripped_regency = upper_regency
    if upper_regency.startswith("KABUPATEN "):
        stripped_regency = upper_regency[10:].strip()
    elif upper_regency.startswith("KOTA "):
        stripped_regency = upper_regency[5:].strip()
    elif upper_regency.startswith("CITY OF "):
        stripped_regency = upper_regency[8:].strip()
        
    if stripped_regency not in ["MAGELANG", "SEMARANG", "TEGAL", "PEKALONGAN"]:
        if regency.upper().startswith("KABUPATEN "):
            regency = regency[10:].strip()
        elif regency.upper().startswith("KOTA "):
            regency = regency[5:].strip()
        elif regency.lower().startswith("city of "):
            regency = regency[8:].strip()

    complete = bool(village and district and regency and province)
    formatted = results[0].get("formatted_address", "") if results else ""
    return (village, district, regency, province), complete, formatted

def google_reverse_geocode(session: requests.Session, lat: float, lng: float, key: str) -> tuple[tuple[str, str, str, str], bool, str, str]:
    params = {"latlng": f"{lat:.8f},{lng:.8f}", "key": key, "language": "id", "region": "id"}
    last_error = ""
    for attempt in range(6):
        try:
            response = session.get(API_URL, params=params, timeout=30)
            response.raise_for_status()
            payload = response.json(); status = payload.get("status", "UNKNOWN_ERROR")
            if status == "OK" and payload.get("results"):
                address, complete, formatted = structured_address(payload["results"])
                return address, complete, formatted, "OK" if complete else "PERLU_VERIFIKASI_KOMPONEN"
            if status not in {"OVER_QUERY_LIMIT", "UNKNOWN_ERROR"}:
                detail = payload.get("error_message", "")
                return ("", "", "", ""), False, "", f"{status}: {detail}".rstrip(": ")
            last_error = status
        except (requests.RequestException, ValueError) as exc:
            last_error = str(exc)
        time.sleep(min(30, 1.5 ** attempt) + random.uniform(0, .4))
    return ("", "", "", ""), False, "", f"GAGAL_SETELAH_RETRY: {last_error}"

def process_google(input_xlsx: Path, output_xlsx: Path, key: str, delay: float = 0.08) -> dict:
    key = key.strip()
    if not key:
        raise ValueError("GOOGLE_MAPS_API_KEY kosong/belum diatur.")
    if not input_xlsx.is_file():
        raise FileNotFoundError(f"File sumber tidak ditemukan: {input_xlsx}")
    if output_xlsx.resolve() == input_xlsx.resolve():
        raise ValueError("Output harus berbeda dari input.")
        
    source = load_workbook(input_xlsx, read_only=True, data_only=True)
    ws = source[source.sheetnames[0]]
    lat_col, lng_col, headers = find_lat_lng_cols(ws)
    out = Workbook(); result = out.active; result.title = "Hasil"; log = out.create_sheet("Log")
    result.append(headers + ["desa", "kecamatan", "kabupaten", "provinsi"])
    log.append(["baris_sumber", "latitude", "longitude", "status", "alamat_google", "catatan"])
    cache: dict[tuple[float, float], tuple[str, bool, str, str]] = {}; session = requests.Session(); counts: defaultdict[str, int] = defaultdict(int)
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        lng_raw, lat_raw = row[lng_col-1], row[lat_col-1]
        try:
            lat, lng = float(lat_raw), float(lng_raw)
            if not (-90 <= lat <= 90 and -180 <= lng <= 180): raise ValueError("di luar rentang koordinat")
        except (TypeError, ValueError) as exc:
            result.append(list(row) + ["", "", "", ""]); log.append([row_no, lat_raw, lng_raw, "KOORDINAT_TIDAK_VALID", "", str(exc)]); counts["KOORDINAT_TIDAK_VALID"] += 1; continue
        coordinate = (round(lat, 8), round(lng, 8))
        if coordinate not in cache:
            cache[coordinate] = google_reverse_geocode(session, lat, lng, key); time.sleep(delay)
        address_tuple, complete, google_address, status = cache[coordinate]
        result.append(list(row) + list(address_tuple)); log.append([row_no, lat, lng, status, google_address, "" if complete else "Periksa hasil Google sebelum digunakan sebagai alamat final."])
        counts[status] += 1
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in (result, log):
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = header_fill; cell.alignment = Alignment(horizontal="center")
    output_xlsx.parent.mkdir(parents=True, exist_ok=True); out.save(output_xlsx)
    return counts
