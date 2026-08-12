import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import shutil
from copy import copy

def clean_durasi(val):
    s = str(val).strip().upper()
    if '<1' in s or '1 MONTH' in s: return '1M'
    if '<2' in s or '2 MONTH' in s: return '2M'
    if '<3' in s or '3 MONTH' in s: return '3M'
    if '4-6' in s: return '4-6M'
    if '>6' in s or '6 MONTH' in s: return '6M'
    return s

def build_w1_occ_map_smart(excel_w1, w1_path, td_filter='ALL'):
    """Build W-1 OCC lookup map strictly from a W-1 table section matching td_filter.
    
    td_filter: 'GREENFIELD', 'BROWNFIELD', or 'ALL'
    - If a matching section is found in W-1's report sheet (e.g. Report per LOP), reads Occ from that section.
    - If report sheet exists in W-1 but does NOT contain the requested table section, returns empty map (Occ W-1 becomes empty '-').
    - If NO report sheet exists in W-1 at all, falls back to raw data sheet 'ODP Golive 2026'.
    """
    w1_map = {}

    sheet_name = None
    for candidate in ['Report per LOP', 'Occ LOP Greenfield', 'Report LOP']:
        if candidate in excel_w1.sheet_names:
            sheet_name = candidate
            break

    if not sheet_name:
        for s in excel_w1.sheet_names:
            if 'LOP' in s.upper() and 'PIVOT' not in s.upper():
                sheet_name = s
                break

    if sheet_name:
        df_w1 = pd.read_excel(w1_path, sheet_name=sheet_name, header=None)
        num_cols = df_w1.shape[1]

        # Case 1: 3-table side-by-side layout in W-1 (num_cols >= 35)
        if num_cols >= 35:
            if td_filter == 'GREENFIELD':
                col_min, col_max = 0, 24
            elif td_filter == 'BROWNFIELD':
                col_min, col_max = 24, 39
            elif td_filter == 'ALL':
                col_min, col_max = 39, num_cols
            else:
                col_min, col_max = 0, num_cols
        else:
            # Case 2: Single-table layout in report sheet (older format W-1 file or single-table report)
            # Greenfield occupies the single table. Brownfield and All Type tables DO NOT exist in W-1 report sheet.
            first_row_str = ' '.join(str(df_w1.iloc[r, c]).strip().upper() for r in range(min(5, len(df_w1))) for c in range(num_cols) if pd.notna(df_w1.iloc[r, c]))
            if td_filter == 'GREENFIELD':
                col_min, col_max = 0, num_cols
            elif td_filter == 'BROWNFIELD' and 'BROWNFIELD' in first_row_str:
                col_min, col_max = 0, num_cols
            elif td_filter == 'ALL' and 'ALL' in first_row_str:
                col_min, col_max = 0, num_cols
            else:
                # Table for td_filter NOT found in W-1 report sheet! Return empty map so Occ W-1 is null/empty!
                print(f"Informasi: Berkas W-1 memiliki sheet laporan '{sheet_name}', tetapi tabel {td_filter} tidak ada di W-1.")
                return {} # DO NOT fall back to raw data sheet when report sheet exists!

        header_row = -1
        col_proyek = -1
        col_branch = -1
        col_wok = -1
        col_durasi = -1
        col_occ = -1

        for r in range(min(15, len(df_w1))):
            for c in range(col_min, min(col_max, num_cols)):
                val = str(df_w1.iloc[r, c]).strip().lower()
                if val in ('nama lop', 'nama proyek', 'proyek', 'ihld lop id', 'nama_lop'):
                    col_proyek = c
                elif val in ('branch', 'telkomsel branch'):
                    col_branch = c
                elif val == 'wok':
                    col_wok = c
                elif any(d in val for d in ['durasi golive', 'cat durasi go live', 'durasi']):
                    col_durasi = c
                elif val == 'occ' or val == 'occupancy':
                    col_occ = c

            if col_proyek != -1 and col_occ != -1:
                header_row = r
                break

        if header_row != -1:
            for r in range(header_row + 1, len(df_w1)):
                row = df_w1.iloc[r]
                p_val = str(row[col_proyek]).strip() if col_proyek != -1 and pd.notna(row[col_proyek]) else ''
                b_val = str(row[col_branch]).strip() if col_branch != -1 and pd.notna(row[col_branch]) else ''
                w_val = str(row[col_wok]).strip() if col_wok != -1 and pd.notna(row[col_wok]) else ''
                d_val = str(row[col_durasi]).strip() if col_durasi != -1 and pd.notna(row[col_durasi]) else ''
                occ_v = row[col_occ] if col_occ != -1 and pd.notna(row[col_occ]) else None

                if not p_val or p_val.lower() in ('nama lop', 'nama proyek', 'none', 'nan', 'row labels', 'grand total', 'type design'):
                    continue

                try:
                    if occ_v is not None and not pd.isna(occ_v):
                        v_float = float(occ_v)
                        dur_c = clean_durasi(d_val)
                        w1_map[(p_val, b_val, w_val, dur_c)] = v_float
                        w1_map[(p_val, dur_c)] = v_float
                        w1_map[p_val] = v_float
                except:
                    pass

            return w1_map
        else:
            print(f"Header row tidak ditemukan untuk tabel {td_filter} di kolom [{col_min}:{col_max}].")
            return {}

    # Option B: Fallback ONLY if NO report sheet exists at all in W-1 (pure raw data file)
    if 'ODP Golive 2026' in excel_w1.sheet_names:
        df_raw = pd.read_excel(w1_path, sheet_name='ODP Golive 2026')
        used_col = 'Used_new_v3' if 'Used_new_v3' in df_raw.columns else 'Used_new_v2'
        df_raw[used_col] = pd.to_numeric(df_raw[used_col], errors='coerce').fillna(0)
        df_raw['Port Terbangun'] = pd.to_numeric(df_raw['Port Terbangun'], errors='coerce').fillna(0)

        if td_filter != 'ALL' and 'Type Design' in df_raw.columns:
            df_raw = df_raw[df_raw['Type Design'].astype(str).str.strip().str.upper() == td_filter]

        agg = df_raw.groupby(['Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live']).agg({
            'Port Terbangun': 'sum',
            used_col: 'sum'
        }).reset_index()

        for _, row in agg.iterrows():
            p_val = str(row['Nama Proyek']).strip()
            b_val = str(row['Telkomsel Branch']).strip()
            w_val = str(row['WOK']).strip()
            d_val = str(row['Cat Durasi Go Live']).strip()
            port = row['Port Terbangun']
            used = row[used_col]
            occ_v = used / port if port > 0 else 0.0

            dur_c = clean_durasi(d_val)
            w1_map[(p_val, b_val, w_val, dur_c)] = occ_v
            w1_map[(p_val, dur_c)] = occ_v
            w1_map[p_val] = occ_v

    return w1_map

def get_w1_occ(row, w1_occ_map):
    if not w1_occ_map:
        return "-"
    proyek = str(row['Nama Proyek']).strip()
    branch = str(row['Telkomsel Branch']).strip()
    wok = str(row['WOK']).strip()
    dur_clean = clean_durasi(row['Cat Durasi Go Live'])

    if (proyek, branch, wok, dur_clean) in w1_occ_map:
        return w1_occ_map[(proyek, branch, wok, dur_clean)]
    if (proyek, dur_clean) in w1_occ_map:
        return w1_occ_map[(proyek, dur_clean)]
    if proyek in w1_occ_map:
        return w1_occ_map[proyek]
    return "-"

def write_table(ws, grouped_df, w1_occ_map, col_start, ref_row=4):
    """Write a LOP table to the worksheet.
    
    col_start: 1-based column number where the table starts:
      - 11 for Greenfield (Cols K-T)
      - 26 for Brownfield (Cols Z-AI)
      - 41 for All Type Design (Cols AO-AX)
    """
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    grouped_df = grouped_df.copy()
    grouped_df['Occ W-1'] = grouped_df.apply(lambda row: get_w1_occ(row, w1_occ_map), axis=1)
    
    def calc_gap(row):
        occ_w1 = row['Occ W-1']
        if isinstance(occ_w1, (int, float)):
            return row['Occ'] - occ_w1
        return "-"
    
    grouped_df['Gap WoW'] = grouped_df.apply(calc_gap, axis=1)
    grouped_df = grouped_df.sort_values(by=['Telkomsel Branch', 'WOK', 'Jumlah_ODP'], ascending=[True, True, False])
    
    start_row = 4
    for _, row in grouped_df.iterrows():
        dest = start_row
        
        # Copy styles from reference row
        if dest > ref_row:
            for i in range(10):
                ref_cell = ws.cell(row=ref_row, column=col_start + i)
                new_cell = ws.cell(row=dest, column=col_start + i)
                if ref_cell.has_style:
                    new_cell.font = copy(ref_cell.font)
                    new_cell.border = copy(ref_cell.border)
                    new_cell.fill = copy(ref_cell.fill)
                    new_cell.number_format = ref_cell.number_format
                    new_cell.protection = copy(ref_cell.protection)
                    new_cell.alignment = copy(ref_cell.alignment)
        
        ws.cell(row=dest, column=col_start + 0).value = row['Nama Proyek']      # K, Z, or AO
        ws.cell(row=dest, column=col_start + 1).value = row['Telkomsel Branch']  # L, AA, or AP
        ws.cell(row=dest, column=col_start + 2).value = row['WOK']              # M, AB, or AQ
        ws.cell(row=dest, column=col_start + 3).value = row['Cat Durasi Go Live'] # N, AC, or AR
        ws.cell(row=dest, column=col_start + 4).value = row['Jumlah_ODP']       # O, AD, or AS
        ws.cell(row=dest, column=col_start + 5).value = row['Port']             # P, AE, or AT
        ws.cell(row=dest, column=col_start + 6).value = row['Used']             # Q, AF, or AU
        
        occ_cell = ws.cell(row=dest, column=col_start + 7)                       # R, AG, or AV
        occ_cell.value = row['Occ']
        if row['Occ'] >= 0.35:
            occ_cell.fill = green_fill
        
        ws.cell(row=dest, column=col_start + 8).value = row['Occ W-1']          # S, AH, or AW
        ws.cell(row=dest, column=col_start + 9).value = row['Gap WoW']          # T, AI, or AX
        start_row += 1

def generate_report_lop(w0_path, w1_path, out_path, template_path, td_filter='GREENFIELD'):
    if not os.path.exists(w0_path):
        raise FileNotFoundError("Berkas W-0 (Raw Data) tidak ditemukan atau telah dihapus.")
    if not os.path.exists(w1_path):
        raise FileNotFoundError("Berkas W-1 (Minggu Lalu) tidak ditemukan atau telah dihapus.")
    if not os.path.exists(template_path):
        raise FileNotFoundError("Berkas Template_LOP_Greenfield.xlsx tidak ditemukan di server.")
        
    temp_w0 = r"temp_lop_w0.xlsx"
    temp_w1 = r"temp_lop_w1.xlsx"
    temp_template = r"temp_lop_template.xlsx"

    try:
        shutil.copy2(w0_path, temp_w0)
        shutil.copy2(w1_path, temp_w1)
        shutil.copy2(template_path, temp_template)
    except Exception as e:
        print(f"Peringatan kopi file: {e}")
        temp_w0, temp_w1, temp_template = w0_path, w1_path, template_path

    # --- VALIDASI BERKAS W-0 ---
    try:
        with pd.ExcelFile(temp_w0) as excel_w0:
            if "ODP Golive 2026" not in excel_w0.sheet_names:
                raise ValueError("Berkas W-0 (Raw Data) tidak memiliki sheet 'ODP Golive 2026'. Harap periksa kembali berkas Anda.")
        df_w0 = pd.read_excel(temp_w0, sheet_name="ODP Golive 2026")
    except ValueError as ve:
        raise ve
    except Exception as e:
        raise ValueError(f"Gagal membaca berkas W-0 (Raw Data): {e}")

    required_cols_w0 = ['Type Design', 'Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live', 'ODP NAME', 'Port Terbangun']
    missing_w0 = [c for c in required_cols_w0 if c not in df_w0.columns]
    if missing_w0:
        raise ValueError(f"Berkas W-0 (Raw Data) tidak memiliki kolom wajib berikut: {', '.join(missing_w0)}. Harap periksa panduan file.")

    used_col = None
    if 'Used_new_v3' in df_w0.columns:
        used_col = 'Used_new_v3'
    elif 'Used_new_v2' in df_w0.columns:
        used_col = 'Used_new_v2'
    else:
        raise ValueError("Berkas W-0 (Raw Data) tidak memiliki kolom 'Used_new_v3' atau 'Used_new_v2'. Harap periksa panduan file.")

    # --- READ W-1 EXCEL ---
    try:
        excel_w1 = pd.ExcelFile(temp_w1)
    except Exception as e:
        raise ValueError(f"Gagal membaca berkas W-1 (Minggu Lalu): {e}")

    # --- PREPARE DATA ---
    df_w0['Type Design Clean'] = df_w0['Type Design'].astype(str).str.strip().str.upper()
    df_w0[used_col] = pd.to_numeric(df_w0[used_col], errors='coerce').fillna(0)
    df_w0.loc[df_w0[used_col] < 0, used_col] = 0
    df_w0['Port Terbangun'] = pd.to_numeric(df_w0['Port Terbangun'], errors='coerce').fillna(0)

    td_clean = td_filter.strip().upper()

    wb = openpyxl.load_workbook(temp_template)
    try:
        sheet_out = None
        if 'Report per LOP' in wb.sheetnames:
            sheet_out = 'Report per LOP'
        elif 'Occ LOP Greenfield' in wb.sheetnames:
            sheet_out = 'Occ LOP Greenfield'
        else:
            sheet_out = wb.sheetnames[0]

        ws = wb[sheet_out]

        if td_clean in ('COMBINED', 'COMBINED_ALL', 'ALL_TABLES', '3_TABLES'):
            # --- GREENFIELD ---
            df_gf = df_w0[df_w0['Type Design Clean'] == 'GREENFIELD'].copy()
            grouped_gf = df_gf.groupby(['Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live']).agg(
                Jumlah_ODP=('ODP NAME', 'count'), Port=('Port Terbangun', 'sum'), Used=(used_col, 'sum')
            ).reset_index()
            grouped_gf['Occ'] = grouped_gf.apply(lambda r: r['Used'] / r['Port'] if r['Port'] > 0 else 0.0, axis=1)

            # --- BROWNFIELD ---
            df_bf = df_w0[df_w0['Type Design Clean'] == 'BROWNFIELD'].copy()
            grouped_bf = df_bf.groupby(['Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live']).agg(
                Jumlah_ODP=('ODP NAME', 'count'), Port=('Port Terbangun', 'sum'), Used=(used_col, 'sum')
            ).reset_index()
            grouped_bf['Occ'] = grouped_bf.apply(lambda r: r['Used'] / r['Port'] if r['Port'] > 0 else 0.0, axis=1)

            # --- ALL TYPE DESIGN ---
            df_all = df_w0.copy()
            grouped_all = df_all.groupby(['Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live']).agg(
                Jumlah_ODP=('ODP NAME', 'count'), Port=('Port Terbangun', 'sum'), Used=(used_col, 'sum')
            ).reset_index()
            grouped_all['Occ'] = grouped_all.apply(lambda r: r['Used'] / r['Port'] if r['Port'] > 0 else 0.0, axis=1)

            w1_occ_map_gf = build_w1_occ_map_smart(excel_w1, temp_w1, 'GREENFIELD')
            w1_occ_map_bf = build_w1_occ_map_smart(excel_w1, temp_w1, 'BROWNFIELD')
            w1_occ_map_all = build_w1_occ_map_smart(excel_w1, temp_w1, 'ALL')

            write_table(ws, grouped_gf, w1_occ_map_gf, col_start=11)
            write_table(ws, grouped_bf, w1_occ_map_bf, col_start=26)
            write_table(ws, grouped_all, w1_occ_map_all, col_start=41)

        else:
            # Single-table mode (Greenfield, Brownfield, or All Type Design)
            if td_clean == 'BROWNFIELD':
                df_sub = df_w0[df_w0['Type Design Clean'] == 'BROWNFIELD'].copy()
                title_text = "LOP Brownfield Golive 2026"
            elif td_clean in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
                df_sub = df_w0.copy()
                title_text = "LOP All Type Design Golive 2026"
            else:
                df_sub = df_w0[df_w0['Type Design Clean'] == 'GREENFIELD'].copy()
                title_text = "LOP Greenfield Golive 2026"

            grouped_sub = df_sub.groupby(['Nama Proyek', 'Telkomsel Branch', 'WOK', 'Cat Durasi Go Live']).agg(
                Jumlah_ODP=('ODP NAME', 'count'), Port=('Port Terbangun', 'sum'), Used=(used_col, 'sum')
            ).reset_index()
            grouped_sub['Occ'] = grouped_sub.apply(lambda r: r['Used'] / r['Port'] if r['Port'] > 0 else 0.0, axis=1)

            w1_occ_map = build_w1_occ_map_smart(excel_w1, temp_w1, td_clean if td_clean != 'GREENFIELD' else 'GREENFIELD')

            # Write single table at cols K-T (col_start=11)
            ws['K1'] = title_text
            write_table(ws, grouped_sub, w1_occ_map, col_start=11)

            # Remove Brownfield & All Type columns from template
            ws.delete_cols(24, 30)

        wb.save(out_path)
    finally:
        try:
            excel_w1.close()
        except:
            pass
        try:
            wb.close()
        except Exception:
            pass
        if os.path.exists(temp_w0) and temp_w0 != w0_path:
            try: os.remove(temp_w0)
            except: pass
        if os.path.exists(temp_w1) and temp_w1 != w1_path:
            try: os.remove(temp_w1)
            except: pass
        if os.path.exists(temp_template) and temp_template != template_path:
            try: os.remove(temp_template)
            except: pass
