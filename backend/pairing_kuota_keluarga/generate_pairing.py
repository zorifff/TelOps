import pandas as pd
import numpy as np
import io
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_msisdn_val(val):
    if pd.isna(val) or val is None or str(val).strip() in ('', 'nan', 'None', '\\N'):
        return ""
    try:
        val_float = float(val)
        if np.isnan(val_float):
            return ""
        return str(int(val_float))
    except (ValueError, TypeError):
        val_str = str(val).strip()
        if val_str.endswith('.0'):
            val_str = val_str[:-2]
        return val_str

def clean_str_val(val):
    if pd.isna(val) or val is None or str(val).strip() in ('', 'nan', 'None', '\\N'):
        return ""
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    return val_str

def generate_pairing_report(input_file_source, mode: str = "option_a") -> bytes:
    """
    Transform Telkomsel One / Kuota Keluarga pairing data from Long (1:M) to Wide (1:1) format
    with TelOps Maroon header styling, auto column width, and grouped child column ordering.
    """
    print(f"=== TelOps Backend: Processing Pairing Kuota Keluarga (Mode: {mode}) ===")
    t0 = time.time()

    # Read input Excel dataframe
    df = pd.read_excel(input_file_source, sheet_name=0)
    print(f"Read {len(df)} rows and {len(df.columns)} columns.")

    # Clean MSISDNs & Key Group Columns
    if 'bb_id' in df.columns:
        df['bb_id'] = df['bb_id'].apply(clean_msisdn_val)
    if 'msisdn_parent' in df.columns:
        df['msisdn_parent'] = df['msisdn_parent'].apply(clean_msisdn_val)
    if 'msisdn_child' in df.columns:
        df['msisdn_child'] = df['msisdn_child'].apply(clean_msisdn_val)

    group_cols = [c for c in ['bb_id', 'msisdn_parent'] if c in df.columns]
    if not group_cols:
        raise ValueError("File Excel tidak memiliki kolom 'bb_id' atau 'msisdn_parent'.")

    # Determine metadata columns (1:1 per group)
    all_possible_metadata = [
        'product_commercial_name', 'activation_date_ih', 'activation_date_parent',
        'city', 'region', 'area', 'cluster', 'sto', 'tsel_id_ih',
        'tsel_id_mobile_parent', 'order_id'
    ]
    metadata_cols = [c for c in all_possible_metadata if c in df.columns]

    # Determine child columns to unstack based on mode
    if mode == "option_b":
        child_cols = ['msisdn_child']
    else:
        child_cols = ['msisdn_child', 'activation_date_child', 'tsel_id_mobile_child']

    child_cols = [c for c in child_cols if c in df.columns]

    # Calculate child sequence index per group
    df['child_seq'] = df.groupby(group_cols).cumcount() + 1
    max_children = df['child_seq'].max() if len(df) > 0 else 0

    # Build unique group dataframe for metadata
    df_groups = df.drop_duplicates(subset=group_cols)[group_cols + metadata_cols].copy()

    # Pivot child columns
    pivoted_dict = {}
    if max_children > 0 and child_cols:
        for c_col in child_cols:
            piv = df.pivot(index=group_cols, columns='child_seq', values=c_col)
            piv.columns = [f"{c_col}{seq}" for seq in piv.columns]
            pivoted_dict[c_col] = piv

        df_children_wide = pd.concat(list(pivoted_dict.values()), axis=1).reset_index()
        df_final = pd.merge(df_groups, df_children_wide, on=group_cols, how='left')
    else:
        df_final = df_groups

    # Clean string formatting across final dataframe
    for col in df_final.columns:
        df_final[col] = df_final[col].apply(clean_str_val)

    # Reorder columns as explicitly requested by User:
    # bb_id, msisdn_parent, msisdn_child1..N, product_commercial_name, activation_date_ih,
    # activation_date_parent, activation_date_child1..N, city, region, area, cluster, sto,
    # tsel_id_ih, tsel_id_mobile_parent, tsel_id_mobile_child1..N, order_id
    ordered_cols = []
    
    # 1. bb_id, msisdn_parent
    for c in ['bb_id', 'msisdn_parent']:
        if c in df_final.columns:
            ordered_cols.append(c)

    # 2. msisdn_child1..N
    for seq in range(1, max_children + 1):
        col_name = f"msisdn_child{seq}"
        if col_name in df_final.columns:
            ordered_cols.append(col_name)

    # 3. product_commercial_name, activation_date_ih, activation_date_parent
    for c in ['product_commercial_name', 'activation_date_ih', 'activation_date_parent']:
        if c in df_final.columns:
            ordered_cols.append(c)

    # 4. activation_date_child1..N
    if mode == "option_a":
        for seq in range(1, max_children + 1):
            col_name = f"activation_date_child{seq}"
            if col_name in df_final.columns:
                ordered_cols.append(col_name)

    # 5. city, region, area, cluster, sto
    for c in ['city', 'region', 'area', 'cluster', 'sto']:
        if c in df_final.columns:
            ordered_cols.append(c)

    # 6. tsel_id_ih, tsel_id_mobile_parent
    for c in ['tsel_id_ih', 'tsel_id_mobile_parent']:
        if c in df_final.columns:
            ordered_cols.append(c)

    # 7. tsel_id_mobile_child1..N
    if mode == "option_a":
        for seq in range(1, max_children + 1):
            col_name = f"tsel_id_mobile_child{seq}"
            if col_name in df_final.columns:
                ordered_cols.append(col_name)

    # 8. order_id
    if 'order_id' in df_final.columns:
        ordered_cols.append('order_id')

    # Add any remaining unlisted columns if present
    for col in df_final.columns:
        if col not in ordered_cols:
            ordered_cols.append(col)

    df_final = df_final[ordered_cols]

    print(f"Transformed to {len(df_final)} unique rows and {len(df_final.columns)} columns.")

    # Pre-calculate column widths vectorized
    col_widths = {}
    for col_idx, col_name in enumerate(df_final.columns, start=1):
        hdr_len = len(str(col_name))
        data_max_len = df_final[col_name].astype(str).str.len().max()
        max_len = max(hdr_len, data_max_len if not pd.isna(data_max_len) else 0)
        col_widths[get_column_letter(col_idx)] = max(int(max_len) + 4, 12)

    # Write & Style with OpenPyXL via ExcelWriter
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='Unstacked_Pairing_Data', index=False)
        ws = writer.sheets['Unstacked_Pairing_Data']

        maroon_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        header_font = Font(name="Poppins", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.fill = maroon_fill
            cell.font = header_font
            cell.alignment = header_align

        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    output_buffer.seek(0)
    print(f"Completed in {time.time() - t0:.2f} seconds.")
    return output_buffer.getvalue()
