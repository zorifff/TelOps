import pandas as pd
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
import os
import shutil
import argparse


woks_in_template = [
    'KEBUMEN', 'MAGELANG TEMANGGUNG', 'BATANG', 'PEMALANG PURBALINGGA', 
    'TEGAL BREBES', 'CILACAP BANYUMAS', 'WONOSOBO BANJARNEGARA', 'DEMAK', 
    'JEPARA KUDUS - PATI', 'SEMARANG 1', 'SEMARANG 2', 'BOYOLALI', 
    'SRAGEN', 'SURAKARTA', 'YOGYA 1', 'YOGYA 2'
]

# Table layout configs (for COMBINED 3-table mode)
TABLE_CONFIGS = [
    {
        'name': 'GREENFIELD',
        'filter': 'GREENFIELD',
        'wok_start': 6,     # first WOK row
        'wok_end': 21,      # last WOK row (inclusive)
        'branch_start': 23, # first branch row
        'branch_end': 28,   # last branch row (inclusive)
        'total_row': 29,    # Jateng DIY row
    },
    {
        'name': 'BROWNFIELD',
        'filter': 'BROWNFIELD',
        'wok_start': 39,
        'wok_end': 54,
        'branch_start': 56,
        'branch_end': 61,
        'total_row': 62,
    },
    {
        'name': 'ALL TYPE DESIGN',
        'filter': 'ALL',  # no filter, take both Greenfield + Brownfield
        'wok_start': 72,
        'wok_end': 87,
        'branch_start': 89,
        'branch_end': 94,
        'total_row': 95,
    },
]

# Single-table config (always uses the Greenfield block position in template)
SINGLE_TABLE_CFG = {
    'wok_start': 6,
    'wok_end': 21,
    'branch_start': 23,
    'branch_end': 28,
    'total_row': 29,
}

# Title text per type design
TITLE_MAP = {
    'GREENFIELD': 'Tracking Occupancy ODP New Golive 2026 - Greenfield',
    'BROWNFIELD': 'Tracking Occupancy ODP New Golive 2026 - Brownfield',
    'ALL': 'Tracking Occupancy ODP New Golive 2026 - All Type Design',
}

SUBTITLE_MAP = {
    'GREENFIELD': 'Occupancy ODP Greenfield 2026',
    'BROWNFIELD': 'Occupancy ODP Brownfield 2026',
    'ALL': 'Occupancy ODP All Type Design 2026',
}

# Branch → WOK index mapping (0-based offset from wok_start)
BRANCH_MAP = [
    ('MAGELANG', (0, 1)),       # KEBUMEN, MAGELANG TEMANGGUNG
    ('PEKALONGAN', (2, 4)),     # BATANG, PEMALANG PURBALINGGA, TEGAL BREBES
    ('PURWOKERTO', (5, 6)),     # CILACAP BANYUMAS, WONOSOBO BANJARNEGARA
    ('SEMARANG', (7, 10)),      # DEMAK, JEPARA KUDUS - PATI, SEMARANG 1, SEMARANG 2
    ('SURAKARTA', (11, 13)),    # BOYOLALI, SRAGEN, SURAKARTA
    ('YOGYAKARTA', (14, 15)),   # YOGYA 1, YOGYA 2
]

# Duration bucket column mapping (Port col, Used col)
BUCKET_COLS = {
    '<1 Month': (3, 4),    # C, D
    '<2 Month': (6, 7),    # F, G
    '<3 Month': (9, 10),   # I, J
    '4-6 Month': (12, 13), # L, M
    '>6 Month': (15, 16),  # O, P
}


def map_bucket(val):
    """Map raw duration category string to standard bucket name."""
    val = str(val).strip().upper()
    if '<1' in val:
        return '<1 Month'
    elif '<2' in val or '2 MONTH' in val:
        return '<2 Month'
    elif '<3' in val or '3 MONTH' in val:
        return '<3 Month'
    elif '4-6' in val:
        return '4-6 Month'
    elif '>6' in val:
        return '>6 Month'
    return val


def find_w1_table_bounds(df_w1_rep, td_filter):
    """
    Find the wok_start and wok_end in df_w1_rep for td_filter by matching sheet headers.
    Returns (wok_start, wok_end) 1-based row numbers, or (None, None) if not found.
    """
    td_clean = td_filter.upper()
    target_keyword = None
    if td_clean == 'GREENFIELD':
        target_keyword = 'GREENFIELD'
    elif td_clean == 'BROWNFIELD':
        target_keyword = 'BROWNFIELD'
    elif td_clean in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
        target_keyword = 'ALL TYPE'

    if not target_keyword:
        return None, None

    for idx, row in df_w1_rep.iterrows():
        cell_val = ""
        if len(row) > 1 and pd.notna(row[1]):
            cell_val += str(row[1]).upper() + " "
        if len(row) > 2 and pd.notna(row[2]):
            cell_val += str(row[2]).upper()

        if 'TRACKING OCCUPANCY' in cell_val and target_keyword in cell_val:
            wok_start = idx + 5
            wok_end = idx + 20
            return wok_start, wok_end

    return None, None


def extract_w1_occ(excel_w1, w1_path, td_filter, wok_start=None, wok_end=None, branch_start=None, branch_end=None, total_row=None):
    """
    Extract OCC W-1 values for a table block.
    Supports reading from previous report sheet ('Report - Occupancy') or raw sheet ('ODP Golive 2026').
    Handles uncomputed Excel formulas gracefully by calculating OCC from raw Port & Used bucket columns.
    """
    w1_occ_wok = {}
    w1_occ_branch = {}

    # Option A: Read from previous Report - Occupancy sheet if present and row range exists
    if 'Report - Occupancy' in excel_w1.sheet_names:
        df_w1_rep = pd.read_excel(w1_path, sheet_name='Report - Occupancy', header=None)

        # Dynamically locate table bounds in W-1 sheet by title matching
        found_start, found_end = find_w1_table_bounds(df_w1_rep, td_filter)
        if found_start is not None:
            wok_start = found_start
            wok_end = found_end

        if wok_start is not None and len(df_w1_rep) >= wok_end:
            wok_port = {}
            wok_used = {}

            for r in range(wok_start, wok_end + 1):
                row = df_w1_rep.iloc[r - 1]
                wok_raw = str(row[1]).strip() if pd.notna(row[1]) else ''
                if not wok_raw or wok_raw.upper() in ('WOK', 'NAN', 'NONE', '') or 'TRACKING' in wok_raw.upper():
                    continue
                wok_name = wok_raw.upper()

                # First try reading cached OCC from Col T (index 19)
                val_t = row[19] if len(row) > 19 else None
                got_t = False
                try:
                    if pd.notnull(val_t) and not pd.isna(val_t):
                        w1_occ_wok[wok_name] = float(val_t)
                        got_t = True
                except:
                    pass

                # Sum raw Port and Used values from duration buckets (Cols C, D, F, G, I, J, L, M, O, P)
                p_sum = sum(float(row[c]) for c in [2, 5, 8, 11, 14] if len(row) > c and pd.notnull(row[c]) and not pd.isna(row[c]))
                u_sum = sum(float(row[c]) for c in [3, 6, 9, 12, 15] if len(row) > c and pd.notnull(row[c]) and not pd.isna(row[c]))

                wok_port[wok_name] = p_sum
                wok_used[wok_name] = u_sum

                if not got_t:
                    w1_occ_wok[wok_name] = u_sum / p_sum if p_sum > 0 else 0.0

            # Calculate Branch and Jateng DIY OCC W-1 from member WOK sums
            for b_name, (w_start, w_end) in BRANCH_MAP:
                target_woks = woks_in_template[w_start:w_end+1]
                bp = sum(wok_port.get(w, 0.0) for w in target_woks)
                bu = sum(wok_used.get(w, 0.0) for w in target_woks)
                w1_occ_branch[b_name] = bu / bp if bp > 0 else 0.0

            tot_p = sum(wok_port.values())
            tot_u = sum(wok_used.values())
            w1_occ_branch['JATENG DIY'] = tot_u / tot_p if tot_p > 0 else 0.0

            if len(w1_occ_wok) > 0:
                return w1_occ_wok, w1_occ_branch
        else:
            print(f"   Informasi: Sheet 'Report - Occupancy' pada W-1 tidak memiliki tabel {td_filter}, OCC W-1 tidak terisi.")
            return {}, {}

    # Option B: Fallback to raw ODP Golive 2026 sheet ONLY if Report - Occupancy sheet is NOT present at all in W-1
    if 'ODP Golive 2026' in excel_w1.sheet_names:
        df_w1_raw = pd.read_excel(w1_path, sheet_name='ODP Golive 2026')
        used_col = 'Used_new_v3' if 'Used_new_v3' in df_w1_raw.columns else 'Used_new_v2'
        df_w1_raw[used_col] = pd.to_numeric(df_w1_raw[used_col], errors='coerce').fillna(0)
        df_w1_raw['Port Terbangun'] = pd.to_numeric(df_w1_raw['Port Terbangun'], errors='coerce').fillna(0)
        df_w1_raw['WOK_Clean'] = df_w1_raw['WOK'].astype(str).str.strip().str.upper()
        df_w1_raw['TD_Clean'] = df_w1_raw['Type Design'].astype(str).str.strip().str.upper()

        if td_filter != 'ALL':
            sub = df_w1_raw[df_w1_raw['TD_Clean'] == td_filter]
        else:
            sub = df_w1_raw

        agg = sub.groupby('WOK_Clean').agg({'Port Terbangun': 'sum', used_col: 'sum'})
        for wok in woks_in_template:
            if wok in agg.index:
                p = agg.loc[wok, 'Port Terbangun']
                u = agg.loc[wok, used_col]
                w1_occ_wok[wok] = u / p if p > 0 else 0.0
            else:
                w1_occ_wok[wok] = 0.0

        for b_name, (w_start, w_end) in BRANCH_MAP:
            woks = woks_in_template[w_start:w_end+1]
            sub_b = agg[agg.index.isin(woks)]
            p_b = sub_b['Port Terbangun'].sum()
            u_b = sub_b[used_col].sum()
            w1_occ_branch[b_name] = u_b / p_b if p_b > 0 else 0.0

        tot_p = agg['Port Terbangun'].sum()
        tot_u = agg[used_col].sum()
        w1_occ_branch['JATENG DIY'] = tot_u / tot_p if tot_p > 0 else 0.0

    return w1_occ_wok, w1_occ_branch


def aggregate_raw_data(df_raw, type_filter):
    """
    Aggregate raw data by WOK and duration bucket.
    type_filter: 'GREENFIELD', 'BROWNFIELD', or 'ALL'
    Returns dict: { WOK_NAME: { bucket: (port_sum, used_sum) } }
    """
    df = df_raw.copy()

    # Filter by Type Design
    if 'Type Design' in df.columns and type_filter != 'ALL':
        df = df[df['Type Design'].astype(str).str.strip().str.upper() == type_filter].copy()

    # Ensure numeric columns
    used_col = 'Used_new_v3' if 'Used_new_v3' in df.columns else 'Used_new_v2'
    df[used_col] = pd.to_numeric(df[used_col], errors='coerce').fillna(0)
    df['Port Terbangun'] = pd.to_numeric(df['Port Terbangun'], errors='coerce').fillna(0)

    # Map duration buckets
    df['Durasi_Clean'] = df['Cat Durasi Go Live'].astype(str).str.strip()
    df['Bucket'] = df['Durasi_Clean'].apply(map_bucket)

    # Aggregate
    agg = df.groupby(['WOK', 'Bucket']).agg({
        'Port Terbangun': 'sum',
        used_col: 'sum'
    }).reset_index()

    w0_data = {}
    for _, row in agg.iterrows():
        wok = str(row['WOK']).strip().upper()
        bucket = row['Bucket']
        if wok not in w0_data:
            w0_data[wok] = {}
        w0_data[wok][bucket] = (int(row['Port Terbangun']), int(row[used_col]))

    return w0_data


def fill_table_block(ws, w0_data, w1_occ_wok, w1_occ_branch, cfg):
    """
    Fill one table block (WOK rows + branch rows + total row) in the worksheet.
    Writes Port/Used raw values, OCC W-1 values, and all Excel formulas.
    """
    wok_start = cfg['wok_start']
    wok_end = cfg['wok_end']
    branch_start = cfg['branch_start']
    total_row = cfg['total_row']

    # ── 1. Fill WOK data rows ──
    for idx, r in enumerate(range(wok_start, wok_end + 1)):
        wok_cell = ws.cell(row=r, column=2)
        wok_raw = str(wok_cell.value).strip() if wok_cell.value else ""
        if not wok_raw or wok_raw == "None":
            continue

        wok_upper = wok_raw.upper()

        # Fill Port and Used for each bucket
        for bucket, (col_p, col_u) in BUCKET_COLS.items():
            port_val = 0
            used_val = 0
            if wok_upper in w0_data and bucket in w0_data[wok_upper]:
                port_val, used_val = w0_data[wok_upper][bucket]
            ws.cell(row=r, column=col_p).value = port_val
            ws.cell(row=r, column=col_u).value = used_val

        # Formulas for WOK row
        ws[f'E{r}'] = f'=IFERROR(D{r}/C{r},"-")'
        ws[f'H{r}'] = f'=IFERROR(G{r}/F{r},"-")'
        ws[f'K{r}'] = f'=IFERROR(J{r}/I{r},"-")'
        ws[f'N{r}'] = f'=IFERROR(M{r}/L{r},0)'
        ws[f'Q{r}'] = f'=IFERROR(P{r}/O{r},0)'
        ws[f'R{r}'] = f'=SUM(C{r},F{r},I{r},L{r},O{r})'
        ws[f'S{r}'] = f'=SUM(D{r},G{r},J{r},M{r},P{r})'
        ws[f'T{r}'] = f'=IFERROR(S{r}/R{r},0)'

        # Fill OCC W-1 (Column U → index 21) and Gap WoW (Column V → index 22)
        if wok_upper in w1_occ_wok:
            ws[f'U{r}'] = w1_occ_wok[wok_upper]
            ws[f'V{r}'] = f'=IF(ISNUMBER(U{r}),T{r}-U{r},"-")'
        else:
            ws[f'U{r}'] = ""
            ws[f'V{r}'] = f'=IF(ISNUMBER(U{r}),T{r}-U{r},"-")'

    # ── 2. Fill Branch rows ──
    for b_idx, (b_name, (w_off_start, w_off_end)) in enumerate(BRANCH_MAP):
        br = branch_start + b_idx
        w_first = wok_start + w_off_start
        w_last = wok_start + w_off_end

        ws[f'C{br}'] = f'=SUM(C{w_first}:C{w_last})'
        ws[f'D{br}'] = f'=SUM(D{w_first}:D{w_last})'
        ws[f'E{br}'] = f'=IFERROR(D{br}/C{br},"-")'

        ws[f'F{br}'] = f'=SUM(F{w_first}:F{w_last})'
        ws[f'G{br}'] = f'=SUM(G{w_first}:G{w_last})'
        ws[f'H{br}'] = f'=IFERROR(G{br}/F{br},"-")'

        ws[f'I{br}'] = f'=SUM(I{w_first}:I{w_last})'
        ws[f'J{br}'] = f'=SUM(J{w_first}:J{w_last})'
        ws[f'K{br}'] = f'=IFERROR(J{br}/I{br},0)'

        ws[f'L{br}'] = f'=SUM(L{w_first}:L{w_last})'
        ws[f'M{br}'] = f'=SUM(M{w_first}:M{w_last})'
        ws[f'N{br}'] = f'=IFERROR(M{br}/L{br},0)'

        ws[f'O{br}'] = f'=SUM(O{w_first}:O{w_last})'
        ws[f'P{br}'] = f'=SUM(P{w_first}:P{w_last})'
        ws[f'Q{br}'] = f'=IFERROR(P{br}/O{br},0)'

        ws[f'R{br}'] = f'=SUM(R{w_first}:R{w_last})'
        ws[f'S{br}'] = f'=SUM(S{w_first}:S{w_last})'
        ws[f'T{br}'] = f'=IFERROR(S{br}/R{br},0)'

        # OCC W-1
        if b_name.upper() in w1_occ_branch:
            ws[f'U{br}'] = w1_occ_branch[b_name.upper()]
            ws[f'V{br}'] = f'=IF(ISNUMBER(U{br}),T{br}-U{br},"-")'
        else:
            ws[f'U{br}'] = ""
            ws[f'V{br}'] = f'=IF(ISNUMBER(U{br}),T{br}-U{br},"-")'

    # ── 3. Fill Total row (Jateng DIY) ──
    tr = total_row
    b_first = branch_start
    b_last = branch_start + len(BRANCH_MAP) - 1

    ws[f'C{tr}'] = f'=SUM(C{b_first}:C{b_last})'
    ws[f'D{tr}'] = f'=SUM(D{b_first}:D{b_last})'
    ws[f'E{tr}'] = f'=IFERROR(D{tr}/C{tr},0)'

    ws[f'F{tr}'] = f'=SUM(F{b_first}:F{b_last})'
    ws[f'G{tr}'] = f'=SUM(G{b_first}:G{b_last})'
    ws[f'H{tr}'] = f'=IFERROR(G{tr}/F{tr},0)'

    ws[f'I{tr}'] = f'=SUM(I{b_first}:I{b_last})'
    ws[f'J{tr}'] = f'=SUM(J{b_first}:J{b_last})'
    ws[f'K{tr}'] = f'=IFERROR(J{tr}/I{tr},0)'

    ws[f'L{tr}'] = f'=SUM(L{b_first}:L{b_last})'
    ws[f'M{tr}'] = f'=SUM(M{b_first}:M{b_last})'
    ws[f'N{tr}'] = f'=IFERROR(M{tr}/L{tr},0)'

    ws[f'O{tr}'] = f'=SUM(O{b_first}:O{b_last})'
    ws[f'P{tr}'] = f'=SUM(P{b_first}:P{b_last})'
    ws[f'Q{tr}'] = f'=IFERROR(P{tr}/O{tr},0)'

    ws[f'R{tr}'] = f'=SUM(R{b_first}:R{b_last})'
    ws[f'S{tr}'] = f'=SUM(S{b_first}:S{b_last})'
    ws[f'T{tr}'] = f'=IFERROR(S{tr}/R{tr},0)'

    # OCC W-1 for Jateng DIY
    if 'JATENG DIY' in w1_occ_branch:
        ws[f'U{tr}'] = w1_occ_branch['JATENG DIY']
        ws[f'V{tr}'] = f'=IF(ISNUMBER(U{tr}),T{tr}-U{tr},"-")'
    else:
        ws[f'U{tr}'] = ""
        ws[f'V{tr}'] = f'=IF(ISNUMBER(U{tr}),T{tr}-U{tr},"-")'


def generate_report(file_w0, file_w1, output_file, template_file,
                    sheet_w1_report, sheet_w0_raw, sheet_w0_template,
                    td_filter='COMBINED'):
    """
    Generate Report ODP.
    td_filter: 'GREENFIELD', 'BROWNFIELD', 'ALL', or 'COMBINED'
    """
    # Copy files to avoid lock issues
    temp_w0 = r"temp_w0.xlsx"
    temp_w1 = r"temp_w1.xlsx"
    temp_template = r"temp_template.xlsx"

    try:
        shutil.copy2(file_w0, temp_w0)
        shutil.copy2(file_w1, temp_w1)
        shutil.copy2(template_file, temp_template)
    except Exception as e:
        print(f"Error mengkopi file (mungkin sedang dibuka di Excel?): {e}")
        return

    # ── Step 1: Read W-1 report excel ──
    print(f"1. Membaca berkas W-1 ({file_w1})...")
    try:
        excel_w1 = pd.ExcelFile(temp_w1)
    except Exception as e:
        print(f"Error membaca W-1: {e}")
        return

    # ── Step 2: Read W-0 raw data ──
    print(f"2. Membaca dan Memproses Raw Data W-0 ({file_w0}) sheet '{sheet_w0_raw}'...")
    try:
        df_raw = pd.read_excel(temp_w0, sheet_name=sheet_w0_raw)

        required_cols = ['WOK', 'Type Design', 'Cat Durasi Go Live', 'Port Terbangun']
        missing = [c for c in required_cols if c not in df_raw.columns]
        if missing:
            print(f"Error: Kolom berikut tidak ditemukan di raw data: {', '.join(missing)}")
            return
    except Exception as e:
        print(f"Error memproses Raw Data: {e}")
        return

    # ── Step 3: Write to template ──
    print(f"3. Menulis ke Master Template '{template_file}' pada sheet '{sheet_w0_template}'...")
    try:
        wb = openpyxl.load_workbook(temp_template)
        if sheet_w0_template not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_w0_template}' tidak ditemukan di {template_file}")
            return

        ws = wb[sheet_w0_template]

        td_clean = td_filter.strip().upper()

        if td_clean in ('COMBINED', 'COMBINED_ALL', 'ALL_TABLES', '3_TABLES'):
            # ── COMBINED MODE: Fill all 3 table blocks ──
            for cfg in TABLE_CONFIGS:
                td_f = cfg['filter']
                print(f"   Memproses tabel {cfg['name']}...")

                w0_data = aggregate_raw_data(df_raw, td_f)
                w1_occ_wok, w1_occ_branch = extract_w1_occ(
                    excel_w1, temp_w1, td_f,
                    cfg['wok_start'], cfg['wok_end'],
                    cfg['branch_start'], cfg['branch_end'], cfg['total_row']
                )
                print(f"     W-1: {len(w1_occ_wok)} WOK, {len(w1_occ_branch)} Branch/Total")
                fill_table_block(ws, w0_data, w1_occ_wok, w1_occ_branch, cfg)

            # Re-apply conditional formatting for all 3 tables
            _apply_color_scale(ws, TABLE_CONFIGS)

        else:
            # ── SINGLE-TABLE MODE ──
            if td_clean == 'BROWNFIELD':
                actual_filter = 'BROWNFIELD'
            elif td_clean in ('ALL', 'ALL TYPE', 'ALL TYPE DESIGN'):
                actual_filter = 'ALL'
            else:
                actual_filter = 'GREENFIELD'

            print(f"   Memproses tabel {actual_filter} (single-table mode)...")

            # Update title and subtitle in the Greenfield block position
            ws.cell(row=2, column=2).value = TITLE_MAP.get(actual_filter, TITLE_MAP['GREENFIELD'])
            ws.cell(row=3, column=3).value = SUBTITLE_MAP.get(actual_filter, SUBTITLE_MAP['GREENFIELD'])

            # Aggregate raw data
            w0_data = aggregate_raw_data(df_raw, actual_filter)

            # Extract W-1 OCC for actual_filter (smart title lookup in W-1)
            target_table_cfg = next((c for c in TABLE_CONFIGS if c['filter'] == actual_filter), TABLE_CONFIGS[0])

            w1_occ_wok, w1_occ_branch = extract_w1_occ(
                excel_w1, temp_w1, actual_filter,
                target_table_cfg['wok_start'], target_table_cfg['wok_end'],
                target_table_cfg['branch_start'], target_table_cfg['branch_end'], target_table_cfg['total_row']
            )
            print(f"     W-1: {len(w1_occ_wok)} WOK, {len(w1_occ_branch)} Branch/Total")

            # Fill the single table block at the top position (SINGLE_TABLE_CFG)
            cfg = SINGLE_TABLE_CFG.copy()
            cfg['filter'] = actual_filter
            fill_table_block(ws, w0_data, w1_occ_wok, w1_occ_branch, cfg)

            # Delete the Brownfield and All Type Design table blocks (rows 30-95)
            ws.delete_rows(30, 66)

            # Re-apply conditional formatting for single table only
            _apply_color_scale(ws, [{'wok_start': 6, 'wok_end': 21, 'branch_start': 23, 'branch_end': 28}])

        # Save output
        wb.save(output_file)
        print(f"Selesai! File berhasil disimpan di: {output_file}")

    except Exception as e:
        print(f"Error menulis ke template: {e}")
    finally:
        # Cleanup temporary files
        try:
            if 'excel_w1' in locals():
                excel_w1.close()
        except:
            pass
        if os.path.exists(temp_w0):
            try: os.remove(temp_w0)
            except: pass
        if os.path.exists(temp_template):
            try: os.remove(temp_template)
            except: pass


def _apply_color_scale(ws, configs):
    """Apply conditional formatting (color scale) to OCC columns for given table configs."""
    rule = ColorScaleRule(
        start_type='min', start_color='FFF8696B',
        mid_type='percentile', mid_value=50, mid_color='FFFCFCFF',
        end_type='max', end_color='FF63BE7B'
    )

    for cfg in configs:
        ws_r = cfg['wok_start']
        we_r = cfg['wok_end']
        bs_r = cfg['branch_start']
        be_r = cfg['branch_end']

        for col in ['E', 'H', 'K', 'N', 'Q']:
            ws.conditional_formatting.add(f'{col}{ws_r}:{col}{we_r}', rule)
            ws.conditional_formatting.add(f'{col}{bs_r}:{col}{be_r}', rule)

        # T:U range
        ws.conditional_formatting.add(f'T{ws_r}:U{we_r}', rule)
        ws.conditional_formatting.add(f'T{bs_r}:U{be_r}', rule)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Report Occupancy ODP")
    parser.add_argument("--w0", required=True, help="Path ke file W-0 (Raw Data terbaru)")
    parser.add_argument("--w1", required=True, help="Path ke file W-1 (Laporan minggu lalu)")
    parser.add_argument("--out", required=True, help="Path untuk menyimpan file output")
    parser.add_argument("--template", default="Template_Report.xlsx", help="Path ke file master template")
    parser.add_argument("--td-filter", default="COMBINED", help="Type Design filter: GREENFIELD, BROWNFIELD, ALL, or COMBINED")

    # Optional arguments for sheet names
    parser.add_argument("--sheet-w1-report", default="Table Report", help="Nama sheet report minggu lalu (default: Table Report)")
    parser.add_argument("--sheet-w0-raw", default="ODP Golive 2026", help="Nama sheet raw data terbaru (default: ODP Golive 2026)")
    parser.add_argument("--sheet-w0-template", default="Report - Occupancy", help="Nama sheet pada file template (default: Report - Occupancy)")

    args = parser.parse_args()

    generate_report(args.w0, args.w1, args.out, args.template,
                    args.sheet_w1_report, args.sheet_w0_raw, args.sheet_w0_template,
                    td_filter=args.td_filter)
