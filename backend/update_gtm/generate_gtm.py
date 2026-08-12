import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import shutil


def clean_str(val):
    if pd.isna(val) or val is None:
        return ""
    return str(val).strip()


def match_column(columns, candidates):
    """
    Dynamically find matching column name from candidates list.
    Supports case-insensitive exact match and substring match.
    """
    cols_clean = {str(c).strip().lower(): c for c in columns}
    for cand in candidates:
        cand_clean = cand.strip().lower()
        if cand_clean in cols_clean:
            return cols_clean[cand_clean]
    for cand in candidates:
        cand_clean = cand.strip().lower()
        for col_clean, orig_col in cols_clean.items():
            if cand_clean in col_clean:
                return orig_col
    return None


def find_header_row_and_sheet(excel_file):
    """
    Find raw data sheet and header row dynamically.
    Returns (sheet_name, header_row_idx, df_raw)
    """
    target_sheet = None
    for s in excel_file.sheet_names:
        if 'ODP' in s.upper() and '2026' in s.upper():
            target_sheet = s
            break
    if not target_sheet:
        for s in excel_file.sheet_names:
            if 'ODP' in s.upper():
                target_sheet = s
                break
    if not target_sheet:
        target_sheet = excel_file.sheet_names[0]

    df_temp = pd.read_excel(excel_file, sheet_name=target_sheet, header=None, nrows=20)
    header_row = 0
    for r in range(len(df_temp)):
        row_vals = [str(x).strip().lower() for x in df_temp.iloc[r] if pd.notna(x)]
        matches = sum(1 for v in row_vals if any(k in v for k in ['odp name', 'nama proyek', 'wok', 'telkomsel branch', 'port terbangun', 'latitude']))
        if matches >= 2:
            header_row = r
            break

    df_raw = pd.read_excel(excel_file, sheet_name=target_sheet, header=header_row)
    return target_sheet, header_row, df_raw


def extract_w1_occ_data(w1_path, mode='all'):
    """
    Extract Branch-level occupancy map and total Jateng DIY occupancy from W-1 file.
    Supports both Report Occupancy output files (sheet 'Report - Occupancy')
    and Raw Data Excel files (sheet 'ODP Golive 2026').
    Filters extraction based on mode: 'greenfield', 'brownfield', or 'all'.
    Returns (w1_branch_map, w1_total_occ)
    """
    w1_branch_map = {}
    w1_total_occ = None
    try:
        excel_w1 = pd.ExcelFile(w1_path)

        # Strategy 1: Report Occupancy file (sheet 'Report - Occupancy')
        if 'Report - Occupancy' in excel_w1.sheet_names:
            wb_w1 = openpyxl.load_workbook(w1_path, data_only=True)
            ws_rep = wb_w1['Report - Occupancy']

            # Target exact Branch row ranges (excluding WOK rows to prevent WOK 'SURAKARTA' collision):
            # Table 1 Greenfield Branch rows: 23..29
            # Table 2 Brownfield Branch rows: 56..62
            # Table 3 All Type Design Branch rows: 89..95
            if mode == 'greenfield':
                row_ranges_to_check = [(23, 30)]
            elif mode == 'brownfield':
                row_ranges_to_check = [(56, 63)]
            else:
                row_ranges_to_check = [(89, 96), (23, 30), (56, 63)]

            def parse_occ_from_range(r_start, r_end):
                b_map = {}
                tot_occ = None

                # Scan headers in 5 preceding rows
                col_occ_w1_idx = None
                col_occ_w0_idx = None

                hdr_scan_start = max(1, r_start - 5)
                for r_hdr in range(hdr_scan_start, r_start):
                    for c_idx in range(1, ws_rep.max_column + 1):
                        cell_str = clean_str(ws_rep.cell(row=r_hdr, column=c_idx).value).upper()
                        if cell_str in ('OCC W-1', 'OCC W1', 'OCC_W1', 'OCC MINGGU LALU'):
                            col_occ_w1_idx = c_idx
                        elif cell_str in ('OCC', 'OCC W-0', 'OCC W0', 'OCC_W0') and col_occ_w0_idx is None:
                            col_occ_w0_idx = c_idx

                if col_occ_w1_idx is None:
                    col_occ_w1_idx = 21  # Default Col U
                if col_occ_w0_idx is None:
                    col_occ_w0_idx = 20  # Default Col T

                for r in range(r_start, min(r_end, ws_rep.max_row + 1)):
                    b_name = clean_str(ws_rep.cell(row=r, column=2).value).upper()
                    if not b_name:
                        continue

                    val_w1 = ws_rep.cell(row=r, column=col_occ_w1_idx).value
                    val_w0 = ws_rep.cell(row=r, column=col_occ_w0_idx).value

                    occ_val = None
                    if isinstance(val_w1, (int, float)) and val_w1 > 0:
                        occ_val = float(val_w1)
                    elif isinstance(val_w0, (int, float)) and val_w0 > 0:
                        occ_val = float(val_w0)

                    if b_name in ('MAGELANG', 'PEKALONGAN', 'PURWOKERTO', 'SEMARANG', 'SURAKARTA', 'YOGYAKARTA'):
                        if occ_val is not None and b_name not in b_map:
                            b_map[b_name] = occ_val
                    elif 'JATENG' in b_name or 'TOTAL' in b_name:
                        if occ_val is not None and tot_occ is None:
                            tot_occ = occ_val

                return b_map, tot_occ

            for r_s, r_e in row_ranges_to_check:
                b_sub, t_sub = parse_occ_from_range(r_s, r_e)
                for k, v in b_sub.items():
                    if k not in w1_branch_map and v is not None:
                        w1_branch_map[k] = v
                if w1_total_occ is None and t_sub is not None:
                    w1_total_occ = t_sub
                if len(w1_branch_map) >= 6 and w1_total_occ is not None:
                    break

        # Strategy 2: Raw data sheet 'ODP Golive 2026' (or any sheet containing raw ODP data)
        if not w1_branch_map or w1_total_occ is None:
            sheet_w1, _, df_w1 = find_header_row_and_sheet(excel_w1)
            col_branch = match_column(df_w1.columns, ['telkomsel branch', 'branch', 'tsel branch'])
            col_wok = match_column(df_w1.columns, ['wok'])
            col_td = match_column(df_w1.columns, ['type design', 'typedesign', 'type_design', 'tipe desain'])

            # Correct W-1 Telkomsel Branch based on WOK mapping if WOK column exists
            wok_to_branch_map = {
                'KEBUMEN': 'MAGELANG',
                'MAGELANG TEMANGGUNG': 'MAGELANG',
                'BATANG': 'PEKALONGAN',
                'PEMALANG PURBALINGGA': 'PEKALONGAN',
                'TEGAL BREBES': 'PEKALONGAN',
                'CILACAP BANYUMAS': 'PURWOKERTO',
                'WONOSOBO BANJARNEGARA': 'PURWOKERTO',
                'DEMAK': 'SEMARANG',
                'JEPARA KUDUS - PATI': 'SEMARANG',
                'SEMARANG 1': 'SEMARANG',
                'SEMARANG 2': 'SEMARANG',
                'BOYOLALI': 'SURAKARTA',
                'SRAGEN': 'SURAKARTA',
                'SURAKARTA': 'SURAKARTA',
                'YOGYA 1': 'YOGYAKARTA',
                'YOGYA 2': 'YOGYAKARTA',
            }

            if col_branch and col_wok:
                def correct_w1_branch(row):
                    w_val = clean_str(row[col_wok]).upper()
                    if w_val in wok_to_branch_map:
                        return wok_to_branch_map[w_val]
                    return clean_str(row[col_branch]).upper()
                df_w1[col_branch] = df_w1.apply(correct_w1_branch, axis=1)

            col_used = match_column(df_w1.columns, ['used_new_v3'])
            if not col_used:
                col_used = match_column(df_w1.columns, ['used_new_v2'])
            if not col_used:
                col_used = match_column(df_w1.columns, ['used_new', 'used'])

            col_port = match_column(df_w1.columns, ['port terbangun', 'port_terbangun', 'total port', 'port', 'total'])

            if col_used and col_port:
                df_w1[col_used] = pd.to_numeric(df_w1[col_used], errors='coerce').fillna(0)
                df_w1[col_port] = pd.to_numeric(df_w1[col_port], errors='coerce').fillna(0)

                # Filter W-1 dataframe by mode if col_td exists
                if col_td and mode in ('greenfield', 'brownfield'):
                    df_w1 = df_w1[df_w1[col_td].astype(str).str.strip().str.lower() == mode]

                if w1_total_occ is None:
                    u_tot = float(df_w1[col_used].sum())
                    p_tot = float(df_w1[col_port].sum())
                    if p_tot > 0:
                        w1_total_occ = u_tot / p_tot

                if col_branch and not w1_branch_map:
                    agg = df_w1.groupby(col_branch).agg({col_port: 'sum', col_used: 'sum'})
                    for branch_val, r_row in agg.iterrows():
                        p = float(r_row[col_port])
                        u = float(r_row[col_used])
                        b_clean = clean_str(branch_val).upper()
                        w1_branch_map[b_clean] = (u / p) if p > 0 else 0.0

    except Exception as e:
        print(f"Informasi ekstraksi W-1 OCC data: {e}")

    return w1_branch_map, w1_total_occ


def generate_gtm_report(w0_path, w1_path, out_path):
    """
    Generate Update GTM Excel Report.
    Reads W-0 (Raw Data) and W-1 (Raw Data / Occupancy Report Minggu Lalu).
    Creates sheet 'GTM Requirement Update' with table 'Tracking Data ODP Golive 2026'.
    Filters and sorts alphabetically by: Telkomsel Branch -> WOK -> Nama Proyek -> ODP Name.
    Calculates exact contiguous row range Occupancies for Branch, WOK, and Proyek.
    Adds summary row 'Jateng DIY' with merged cells L:O and total Gap WoW.
    """
    temp_w0 = r"temp_gtm_w0.xlsx"
    temp_w1 = r"temp_gtm_w1.xlsx"

    try:
        shutil.copy2(w0_path, temp_w0)
        shutil.copy2(w1_path, temp_w1)
    except Exception:
        temp_w0, temp_w1 = w0_path, w1_path

    try:
        excel_w0 = pd.ExcelFile(temp_w0)
        sheet_w0, header_row_w0, df_raw = find_header_row_and_sheet(excel_w0)

        # Dynamic column matching for W-0
        col_branch = match_column(df_raw.columns, ['telkomsel branch', 'branch', 'tsel branch'])
        col_wok = match_column(df_raw.columns, ['wok'])
        col_proyek = match_column(df_raw.columns, ['nama proyek', 'proyek', 'project name', 'nama_proyek'])
        col_odp = match_column(df_raw.columns, ['odp name', 'odp_name', 'nama odp', 'odpname', 'odp'])
        col_lat = match_column(df_raw.columns, ['latitude', 'lat'])
        col_long = match_column(df_raw.columns, ['longitude', 'long', 'lng'])
        col_occ2 = match_column(df_raw.columns, ['occ 2', 'occ2', 'occ_2', 'occupancy 2'])
        col_td = match_column(df_raw.columns, ['type design', 'typedesign', 'type_design', 'tipe desain'])

        # Determine W-0 Type Design mode ('greenfield', 'brownfield', or 'all')
        w0_mode = 'all'
        if col_td:
            td_vals = set(clean_str(x).lower() for x in df_raw[col_td].dropna() if clean_str(x))
            if td_vals == {'greenfield'}:
                w0_mode = 'greenfield'
            elif td_vals == {'brownfield'}:
                w0_mode = 'brownfield'

        print(f"Detected W-0 Type Design Mode: [{w0_mode.upper()}]")

        # Prioritize Used_new_v3 -> Used_new_v2 -> Used_new -> Used
        col_used = match_column(df_raw.columns, ['used_new_v3'])
        if not col_used:
            col_used = match_column(df_raw.columns, ['used_new_v2'])
        if not col_used:
            col_used = match_column(df_raw.columns, ['used_new', 'used'])

        col_total = match_column(df_raw.columns, ['port terbangun', 'port_terbangun', 'total port', 'port', 'total'])

        # Validate required columns
        missing = []
        if not col_odp: missing.append("ODP NAME")
        if not col_branch: missing.append("Telkomsel Branch")
        if missing:
            raise ValueError(f"Kolom wajib berikut tidak ditemukan di file W-0 (Raw Data): {', '.join(missing)}")

        # Standard WOK to Branch mapping dictionary
        wok_to_branch_map = {
            'KEBUMEN': 'MAGELANG',
            'MAGELANG TEMANGGUNG': 'MAGELANG',
            'BATANG': 'PEKALONGAN',
            'PEMALANG PURBALINGGA': 'PEKALONGAN',
            'TEGAL BREBES': 'PEKALONGAN',
            'CILACAP BANYUMAS': 'PURWOKERTO',
            'WONOSOBO BANJARNEGARA': 'PURWOKERTO',
            'DEMAK': 'SEMARANG',
            'JEPARA KUDUS - PATI': 'SEMARANG',
            'SEMARANG 1': 'SEMARANG',
            'SEMARANG 2': 'SEMARANG',
            'BOYOLALI': 'SURAKARTA',
            'SRAGEN': 'SURAKARTA',
            'SURAKARTA': 'SURAKARTA',
            'YOGYA 1': 'YOGYAKARTA',
            'YOGYA 2': 'YOGYAKARTA',
        }

        # Correct Telkomsel Branch based on WOK mapping if WOK column exists
        if col_branch and col_wok:
            def correct_branch_val(row):
                w_val = clean_str(row[col_wok]).upper()
                if w_val in wok_to_branch_map:
                    return wok_to_branch_map[w_val]
                return clean_str(row[col_branch]).upper()
            df_raw[col_branch] = df_raw.apply(correct_branch_val, axis=1)

        # Sort dataframe alphabetically: Branch -> WOK -> Nama Proyek -> ODP Name
        sort_cols = [c for c in [col_branch, col_wok, col_proyek, col_odp] if c]
        if sort_cols:
            df_raw = df_raw.sort_values(by=sort_cols, ascending=True).reset_index(drop=True)

        # Compute contiguous row ranges for Branch, WOK, and Proyek
        start_excel_row = 5
        row_count = len(df_raw)
        last_data_row = start_excel_row + row_count - 1

        # 1. Branch contiguous ranges
        branch_range_map = {}  # row_idx -> (start_row, end_row)
        curr_b = None
        s_r = start_excel_row
        for idx, row in df_raw.iterrows():
            b_val = clean_str(row[col_branch]).upper() if col_branch else ""
            excel_r = start_excel_row + idx
            if b_val != curr_b:
                if curr_b is not None:
                    e_r = excel_r - 1
                    for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                        branch_range_map[r_i] = (s_r, e_r)
                curr_b = b_val
                s_r = excel_r
        if curr_b is not None:
            e_r = last_data_row
            for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                branch_range_map[r_i] = (s_r, e_r)

        # 2. WOK contiguous ranges
        wok_range_map = {}  # row_idx -> (start_row, end_row)
        curr_w = None
        s_r = start_excel_row
        for idx, row in df_raw.iterrows():
            b_val = clean_str(row[col_branch]).upper() if col_branch else ""
            w_val = clean_str(row[col_wok]).upper() if col_wok else ""
            bw_key = f"{b_val}||{w_val}"
            excel_r = start_excel_row + idx
            if bw_key != curr_w:
                if curr_w is not None:
                    e_r = excel_r - 1
                    for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                        wok_range_map[r_i] = (s_r, e_r)
                curr_w = bw_key
                s_r = excel_r
        if curr_w is not None:
            e_r = last_data_row
            for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                wok_range_map[r_i] = (s_r, e_r)

        # 3. Proyek contiguous ranges
        proyek_range_map = {}  # row_idx -> (start_row, end_row)
        curr_p = None
        s_r = start_excel_row
        for idx, row in df_raw.iterrows():
            b_val = clean_str(row[col_branch]).upper() if col_branch else ""
            w_val = clean_str(row[col_wok]).upper() if col_wok else ""
            p_val = clean_str(row[col_proyek]).upper() if col_proyek else ""
            bwp_key = f"{b_val}||{w_val}||{p_val}"
            excel_r = start_excel_row + idx
            if bwp_key != curr_p:
                if curr_p is not None:
                    e_r = excel_r - 1
                    for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                        proyek_range_map[r_i] = (s_r, e_r)
                curr_p = bwp_key
                s_r = excel_r
        if curr_p is not None:
            e_r = last_data_row
            for r_i in range(s_r - start_excel_row, e_r - start_excel_row + 1):
                proyek_range_map[r_i] = (s_r, e_r)

        # Extract W-1 Occupancy Data matching W-0 Type Design mode
        w1_branch_map, w1_total_occ = extract_w1_occ_data(temp_w1, mode=w0_mode)
        print(f"Berhasil memuat data W-1 ({w0_mode.upper()}): {len(w1_branch_map)} Branch map, Total Occ W-1={w1_total_occ}")

        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GTM Requirement Update"
        ws.views.sheetView[0].showGridLines = True

        # Write Table Title
        ws.cell(row=2, column=1).value = "Tracking Data ODP Golive 2026"
        ws.cell(row=2, column=1).font = Font(name="Aptos Narrow", size=14, bold=True, color="001A3F")

        # Table Headers (16 Columns in exact requested order)
        headers = [
            "Telkomsel Branch", "WOK", "Nama Proyek", "ODP Name",
            "Latitude", "Longitude", "OCC 2", "Type Design",
            "Used", "Available", "Total",
            "Occ Branch", "Occ WOK", "Occ Proyek", "Occ ODP", "Gap WoW"
        ]

        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(name="Aptos Narrow", size=9, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[4].height = 24

        # Styling for data rows (Aptos Narrow 9pt)
        data_font = Font(name="Aptos Narrow", size=9)
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        for idx, row in df_raw.iterrows():
            r = start_excel_row + idx

            branch_val = clean_str(row[col_branch]) if col_branch else ""
            wok_val = clean_str(row[col_wok]) if col_wok else ""
            proyek_val = clean_str(row[col_proyek]) if col_proyek else ""
            odp_val = clean_str(row[col_odp]) if col_odp else ""
            lat_val = row[col_lat] if col_lat and pd.notna(row[col_lat]) else ""
            long_val = row[col_long] if col_long and pd.notna(row[col_long]) else ""
            occ2_val = clean_str(row[col_occ2]) if col_occ2 else ""
            td_val = clean_str(row[col_td]) if col_td else ""

            try:
                used_num = int(float(row[col_used])) if col_used and pd.notna(row[col_used]) else 0
            except Exception:
                used_num = 0

            try:
                total_num = int(float(row[col_total])) if col_total and pd.notna(row[col_total]) else 0
            except Exception:
                total_num = 0

            # Get range boundaries for formulas
            b_s, b_e = branch_range_map.get(idx, (r, r))
            w_s, w_e = wok_range_map.get(idx, (r, r))
            p_s, p_e = proyek_range_map.get(idx, (r, r))

            # Branch W-1 OCC
            w1_occ_b = w1_branch_map.get(branch_val.upper(), None)

            # Columns 1-8
            c1 = ws.cell(row=r, column=1, value=branch_val); c1.alignment = align_center
            c2 = ws.cell(row=r, column=2, value=wok_val); c2.alignment = align_center
            c3 = ws.cell(row=r, column=3, value=proyek_val); c3.alignment = align_left
            c4 = ws.cell(row=r, column=4, value=odp_val); c4.alignment = align_left
            c5 = ws.cell(row=r, column=5, value=lat_val); c5.alignment = align_center
            c6 = ws.cell(row=r, column=6, value=long_val); c6.alignment = align_center
            c7 = ws.cell(row=r, column=7, value=occ2_val); c7.alignment = align_center
            c8 = ws.cell(row=r, column=8, value=td_val); c8.alignment = align_center

            # Columns 9-11 (Used, Available, Total)
            c9 = ws.cell(row=r, column=9, value=used_num); c9.alignment = align_right; c9.number_format = '#,##0'
            c10 = ws.cell(row=r, column=10, value=f"=K{r}-I{r}"); c10.alignment = align_right; c10.number_format = '#,##0'
            c11 = ws.cell(row=r, column=11, value=total_num); c11.alignment = align_right; c11.number_format = '#,##0'

            # Column 12: Occ Branch = SUM(I$b_s:I$b_e) / SUM(K$b_s:K$b_e)
            c12 = ws.cell(row=r, column=12, value=f'=IF(SUM(K${b_s}:K${b_e})>0, SUM(I${b_s}:I${b_e})/SUM(K${b_s}:K${b_e}), "-")')
            c12.alignment = align_right; c12.number_format = '0.00%'

            # Column 13: Occ WOK = SUM(I$w_s:I$w_e) / SUM(K$w_s:K$w_e)
            c13 = ws.cell(row=r, column=13, value=f'=IF(SUM(K${w_s}:K${w_e})>0, SUM(I${w_s}:I${w_e})/SUM(K${w_s}:K${w_e}), "-")')
            c13.alignment = align_right; c13.number_format = '0.00%'

            # Column 14: Occ Proyek = SUM(I$p_s:I$p_e) / SUM(K$p_s:K$p_e)
            c14 = ws.cell(row=r, column=14, value=f'=IF(SUM(K${p_s}:K${p_e})>0, SUM(I${p_s}:I${p_e})/SUM(K${p_s}:K${p_e}), "-")')
            c14.alignment = align_right; c14.number_format = '0.00%'

            # Column 15: Occ ODP = Used / Total
            c15 = ws.cell(row=r, column=15, value=f'=IF(K{r}>0, I{r}/K{r}, "-")')
            c15.alignment = align_right; c15.number_format = '0.00%'

            # Column 16: Gap WoW = Occ Branch W0 - Occ Branch W1
            c16 = ws.cell(row=r, column=16)
            if w1_occ_b is not None and isinstance(w1_occ_b, (int, float)):
                c16.value = f'=IF(ISNUMBER(L{r}), L{r} - {w1_occ_b}, "-")'
                c16.number_format = '0.00%'
            else:
                c16.value = "-"
            c16.alignment = align_right

            for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16]:
                cell.font = data_font
                cell.border = thin_border

        # Dynamic Summary Row "Jateng DIY" at the bottom
        sum_row = last_data_row + 1
        ws.row_dimensions[sum_row].height = 24

        summary_font = Font(name="Aptos Narrow", size=9, bold=True, color="FFFFFF")
        align_center_bold = Alignment(horizontal="center", vertical="center")
        align_right_bold = Alignment(horizontal="right", vertical="center")

        # 1. Merge Columns A to H (1 to 8) for "Jateng DIY" title
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=8)
        c_title = ws.cell(row=sum_row, column=1, value="Jateng DIY")
        c_title.alignment = align_center_bold

        # 2. Summary Row Columns 9-11 (Used, Available, Total SUM)
        c_used = ws.cell(row=sum_row, column=9, value=f"=SUM(I5:I{last_data_row})")
        c_used.alignment = align_right_bold; c_used.number_format = '#,##0'

        c_avail = ws.cell(row=sum_row, column=10, value=f"=K{sum_row}-I{sum_row}")
        c_avail.alignment = align_right_bold; c_avail.number_format = '#,##0'

        c_tot = ws.cell(row=sum_row, column=11, value=f"=SUM(K5:K{last_data_row})")
        c_tot.alignment = align_right_bold; c_tot.number_format = '#,##0'

        # 3. Merge Columns L to O (12 to 15) for Total Regional Occupancy (Occ Branch to Occ ODP)
        ws.merge_cells(start_row=sum_row, start_column=12, end_row=sum_row, end_column=15)
        c_occ_total = ws.cell(row=sum_row, column=12, value=f'=IF(K{sum_row}>0, I{sum_row}/K{sum_row}, "-")')
        c_occ_total.alignment = align_center_bold; c_occ_total.number_format = '0.00%'

        # 4. Summary Row Column 16: Total Regional Gap WoW = Occ Total W0 - Occ Total W1
        c_gap = ws.cell(row=sum_row, column=16)
        if w1_total_occ is not None and isinstance(w1_total_occ, (int, float)):
            c_gap.value = f'=IF(ISNUMBER(L{sum_row}), L{sum_row} - {w1_total_occ}, "-")'
            c_gap.number_format = '0.00%'
        else:
            c_gap.value = "-"
        c_gap.alignment = align_right_bold

        # Apply Header Fill & White Bold Font & Borders to all cells in Summary Row (Columns 1 to 16)
        for col_i in range(1, 17):
            cell = ws.cell(row=sum_row, column=col_i)
            cell.fill = header_fill
            cell.font = summary_font
            cell.border = thin_border

        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in (2, sum_row):
                    continue  # ignore title & merged summary row text length
                v_str = str(cell.value or '')
                if len(v_str) > max_len:
                    max_len = len(v_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(out_path)
        print(f"File Update GTM berhasil disimpan ke: {out_path}")
    finally:
        for p in [temp_w0, temp_w1]:
            if os.path.exists(p) and p not in (w0_path, w1_path):
                try:
                    os.remove(p)
                except Exception:
                    pass
