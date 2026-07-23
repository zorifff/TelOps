import openpyxl

files = ['backend/automasi_report_odp/Template_Report.xlsx', 'backend/automasi_report_black_odp/Template_Report_Black.xlsx']
for f in files:
    wb = openpyxl.load_workbook(f)
    for ws in wb.worksheets:
        for r in range(1,100):
            for c in range(1,100):
                v = ws.cell(r,c).value
                if v and isinstance(v, str):
                    if '20 Juli 2026' in v:
                        print(f"Found '20 Juli 2026' in {f}, sheet {ws.title}, cell {r},{c}: {v}")
                        ws.cell(r,c).value = v.replace('20 Juli 2026', '').strip()
                        print(f"  Replaced with: {ws.cell(r,c).value}")
    wb.save(f)
    print(f"Saved {f}")
