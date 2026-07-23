import openpyxl
from openpyxl.utils import get_column_letter

template_path = 'backend/automasi_report_lop/Template_LOP_Greenfield.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb['Occ LOP Greenfield']

# 1. We want to move the headers and formatting from columns K-T (11-20) to A-J (1-10)
# We will copy the cell values and formatting (font, fill, border, alignment) for rows 1 to 3 (headers)
import copy

for r in range(1, 4):
    for c in range(1, 11):
        src_cell = ws.cell(row=r, column=c+10) # 11 to 20
        dst_cell = ws.cell(row=r, column=c)    # 1 to 10
        dst_cell.value = src_cell.value
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

# Clear the old columns (Pivot and old headers and everything else)
for r in range(1, 1000):
    for c in range(11, 30):
        ws.cell(row=r, column=c).value = None

wb.save(template_path)
print("Shifted LOP headers to the left.")
